import functools
import io
from io import BytesIO
import numpy as np
import openpyxl
import pandas as pd
import streamlit
from openpyxl.utils.dataframe import dataframe_to_rows
import zipfile
import re
from typing import Dict, Union,List

'''
# 2024.07.08更新：对多个函数进行优化。
# 2024.11.04更新：对整个模块进行重构，增强Andf功能，优化相关功能函数。
# 2024.11.24更新：完成als重构，重命名为al模块。
# 2025.02.01更新：添加了get_cls_score()方法，实现了班级分析-分数段功能；给funs_fd()添加rev参数。
'''

class Andf:
    def __init__(self, df):
        # 返回原始的df表。
        self.__df = df

        # 声明标准学科名列表
        self.__sbj = ["语文", "数学", "英语", "物理", "化学", "生物", "政治", "历史", "地理"]
        # 确定df表对应的[学科名]列表:__sbj_lst.
        self.__sbj_lst = list(set(self.__sbj) & set(self.__df.columns))
        # 确定df表对应的{学科名：序号}字典：__sbj_dic。
        self.__sbj_dic = {val: idx + 1 for idx, val in enumerate(self.__sbj)}
        # 对学科列表进行排序，排序规则为：字典中key对应的值。
        self.__sbj_lst.sort(key=lambda x: self.__sbj_dic[x])

        # 增加总分列
        self.__df["总分"] = df.loc[:, self.__sbj_lst].sum(axis=1)
        # 增加班次列
        self.__df["班次"] = df.groupby("班级")["总分"].rank(ascending=False, method="min")
        # 增加级次列
        self.__df["级次"] = df["总分"].rank(axis=0, ascending=False, method="min")
    def get_sbj_lst(self):
        """
        返回值：学科列表。标准学科名称对应的列表。
        """
        # 对学科列表进行排序，排序规则为：字典中key对应的值。
        return self.__sbj_lst
    def get_sbj_dic(self):
        """
        返回值：学科名：排序的字典。标准学科名称对应的字典("语文":1，"数学":2，"英语":3，"物理":4，"化学":5，"生物":6，"政治":7，"历史":8，"地理":9)
        """
        return dict((key, self.__sbj_dic[key]) for key in self.__sbj_lst)
    def get_all(self):
        """
        返回导入后，添加总分、班次、校次列后的DF表。
        :return: 添加总分、班次、校次列后的DF表。
        """
        return self.__df
    def get_df(self, columns=None, max_class_rank=60):
        """
        返回值：df表。
        :param columns: 列名，列表。
        :param max_class_rank: 班级数，整数。默认40。
        :return: df表。
        """
        if columns is None:
            columns = ["班级", "学号", "姓名"]
        df = self.__df[self.__df["班次"] <= max_class_rank]
        return df.loc[:, columns]
    def get_mc(self, max_class_rank=40,combine_ranks=1):
        """
        返回班次满足小于max_class_rank的df表.当combine_ranks=1合并班次与总分名次为一个元组
        :param max_class_rank: 班级名次，整数。默认40,取班次<=40的数据。
        :param combine_ranks: 整数1或0。0,各科独立排名。默认1,学科排名与总分排名组合为元组.
        :return: df表。班次满足小于max_class_rank的df表.当combine_ranks=1合并班次与总分名次为一个元组
        """
        df = self.__df[self.__df["班次"] <= max_class_rank]
        df_mc = df.loc[:, self.__sbj_lst + ['总分']].rank(axis=0, method="min", ascending=False)
        df_mc = df_mc.astype(object)    # 将数据转换为object类型，以便后续操作将每个元素转换为元组。

        if combine_ranks == 1:
            for i in range(len(self.__sbj_lst)):
                # 将多列的名次数据与总分名次列合并为多个元组列....
                df_mc.iloc[:, i] = df_mc.iloc[:, [i, len(self.__sbj_lst)]].apply(tuple, axis=1)
        return pd.concat([df[["班级", "学号", "姓名"]], df_mc], axis=1)
    def get_fsd(self,dic_val_sbj,thresh_score,max_class_rank=40):
        """
        生成分数段报表.
        :param dic_val_sbj: 阈值列表.如:[0,36,72,96,120]
        :param thresh_score: 学科名次段阈值列表对应积分.如:[10, 9, 2, 1, 0]
        :param max_class_rank: 最大班次,整数.即班级最大参评人数
        :return: 分数段报表.
        """
        # 使用字典推导式转换为{学科:总分}字典
        dic = {value: key for key, values in dic_val_sbj.items() for value in values}
        # 过滤未考学科的键值对。
        dic = {key: value for key, value in dic.items() if key in self.__sbj_lst}
        # 转列表,按标准学科名排序
        lst = sorted(list(dic.items()), key=lambda x: self.get_sbj_dic()[x[0]])
        # 转字典
        dic = dict(lst)
        # res =len(list(dic.values())[0])


        # 定义{学科:[分数段函数组]}字典
        dic_FAN = {key: functools.partial(funs_fd, thresh=dic[key])()  for key, val in dic.items()}

        # 获取需进行两率一平的df表,要求最大班次小于max_class_rank,列索引中人班级与学科名.
        df = self.get_df(columns=["班级"] + self.get_sbj_lst(), max_class_rank=max_class_rank)

        # 按班级分组后，执行分数段函数组。生成一个报表.
        df_fsd = df.groupby("班级").agg(dic_FAN)
        # 依据第0列索引，分割数据为多个df表。
        df_fsd = df_split(df_fsd)

        if thresh_score !=0:
            for dff in df_fsd:
                # 添加参评人数列
                dff.insert(loc=0, column=('统计', "参评人数"),  value=dff.iloc[:, 0:(len(list(dic.values())[0]))].apply(sum, axis=1))
                # 添加积分列
                dff.insert(loc=1, column=('统计', "积分"), value=dff.iloc[:, 1:(len(thresh_score)+1)].apply(functools.partial(np.dot, b=thresh_score), axis=1))
                # 添加排名列.
                dff.insert(loc=2, column=('统计', "排名"), value=dff[("统计", "积分")].rank(axis=0, ascending=False, method="min"))

        return df_fsd
    def get_sdb(self,thresh,thresh_score=[10, 9, 2, 1, 0],max_class_rank=40,max_total_rank=200):
        """
        生成双达标报表.
        :param thresh: 阈值列表.如:[0,200,260,300]
        :param thresh_score: 学科名次段阈值列表对应积分.如:[10, 9, 2, 1, 0]
        :param max_class_rank: 最大班次,整数.即班级最大参评人数
        :param max_total_rank: 最大校次.即最大校次,如:260
        :return: 双达标报表.
        """

        # 获取班级名次（学科名次，总分名次）表。
        df_mc = self.get_mc(max_class_rank=max_class_rank, combine_ranks = 1)

        # 获取学科双达标函数列表
        mcd_funs = functools.partial(funs_mcd_tup, thresh=thresh, max_total_rank=max_total_rank)

        # 创建字典，添加键值对为{学科名：单达标函数},以计算单达标人数.
        dic_mcd_funs = {str(i): mcd_funs() for i in self.__sbj_lst}

        # 使用聚合函数，计算学科双达标人数，并返回df表。
        df_SDB = df_mc.groupby("班级").agg(dic_mcd_funs)

        # 依据第0列索引，分割数据为多个df表。
        dfs_xk = df_split(df_SDB)

        if thresh_score !=0:
            for dff in dfs_xk:
                # 添加参评人数列
                dff.insert(loc=0, column=('统计', "参评人数"), value=dff.iloc[:, 0:(len(thresh_score))].apply(sum, axis=1))
                # 添加积分列
                dff.insert(loc=1, column=('统计', "积分"), value=dff.iloc[:, 1:(len(thresh_score)+1)].apply(functools.partial(np.dot, b=thresh_score), axis=1))
                # 添加排名列.
                dff.insert(loc=2, column=('统计', "排名"), value=dff[("统计", "积分")].rank(axis=0, ascending=False, method="min"))
        return dfs_xk
    def get_lv(self,dic_val_sbj,calcu=1,max_class_rank=40):
        """
        生成两率一平报表.
        :param dic_val_sbj: {学科：[阈值]}字典。如:{120: ["语文", "数学", "英语"], 70: ["物理", "政治"], 50: ["化学", "生物", "历史", "地理"]}
        :param calcu: 0或1，默认1。0，添加。1，添加分析列。
        :param max_class_rank: 最大班次次，整数。默认40,取班次<=40的数据。
        :return: df表。
        """

        # 使用字典推导式转换为{学科:总分}字典
        dic = {value: key for key, values in dic_val_sbj.items() for value in values}
        # 过滤未考学科的键值对。
        dic = {key: value for key, value in dic.items() if key in self.__sbj_lst}
        # 转列表,按标准学科名排序
        lst = sorted(list(dic.items()), key=lambda x: self.get_sbj_dic()[x[0]])
        # 转字典
        dic = dict(lst)

        # 定义{学科:[两率一平函数组]}字典
        dic_FAN = {key: functools.partial(funs_lv,thres=[val*0.6,val*0.8, val])()+["mean"] for key,val in dic.items()}

        # 获取需进行两率一平的df表,要求最大班次小于max_class_rank,列索引中人班级与学科名.
        df = self.get_df(columns=["班级"]+self.get_sbj_lst(),max_class_rank=max_class_rank)

        # 按班级分组后，执行两率一平计算。生成一个报表.
        df_lv = df.groupby("班级").agg(dic_FAN)

        # 依据第0列索引，分割数据为多个df表。
        dfs_lv = df_split(df_lv)

        if calcu != 0:
            for dff in dfs_lv:
                # 添加参评人数列
                dff.insert(loc=0, column=('统计', "参评人数"),value=max_class_rank)
                # 添加排名列.
                dff.insert(loc=1, column=('统计', "排名"), value=dff.iloc[:,5].rank(axis=0, ascending=False, method="min"))
        # # 对列表按标准学科名称排序
        # dfs_lv = sorted(dfs_lv, key=lambda x: (x.columns[4][0]))
        return dfs_lv
    def get_cls(self,thresh= [0,100,200,300],thresh_score = [4,3,2,1],max_class_rank=60,rev=0):
        """
        获取班级分析报表-各次段统计。
        :param thresh: 名次段阈值列表，列表。默认[0,100,200,300]。
        :param thresh_score: 名次段积分列表，列表。默认[4,3,2,1]。
        :param max_class_rank: 最大班级次，整数。默认60。
        :return: df表。班级分析报表。
        """

        # 定义偏函数，设置阈值列表，若为0，则默认为[0,200,400]。cumu:0为不累计，1为累计。
        funs_cls_0 = functools.partial(funs_fd, thresh=thresh,cumu=0,mode=1,inf=1,rev=rev)
        funs_cls_1 = functools.partial(funs_fd, thresh=thresh,cumu=1,mode=1,inf=1,rev=rev)

        # 获取班级名次表["班级"，"级次"]
        df_cls = self.get_df(columns=["班级" ,"级次"],max_class_rank=max_class_rank)
        # 用班级分组，用总分应用名次段函数,count函数用于统计数据总个数。
        df_cls = df_cls.groupby("班级")["级次"].agg(funs_cls_0()+funs_cls_1())

        if thresh_score !=0:
            # 设置偏函数，为点积函数设置b值为默认值.
            fun_jf = functools.partial(np.dot, b=thresh_score)  # 设置偏函数,预先设置qz的值为JF_FSD_SDB.
            # 插入参评人数列
            df_cls.insert(loc=0, column="参评人数", value=df_cls.iloc[:, 0:(len(thresh))].apply(sum, axis=1))
            # 添加积分列
            df_cls.insert(loc=1, column="积分",value=df_cls.iloc[:, 1:(len(thresh_score) + 1)].apply(fun_jf, axis=1))
            # 添加排名列.
            df_cls.insert(loc=2, column="排名",value=df_cls["积分"].rank(axis=0, ascending=False, method="min"))

        return df_cls
    def get_cls_score(self,thresh= [300,400,500,550],thresh_score = [1,2,3,4],max_class_rank=60,rev=0):
        """
        获取班级分析报表--分数段统计。
        :param thresh: 分数段阈值列表，列表。默认[550,500,450,400]。
        :param thresh_score: 名次段积分列表，列表。默认[4,3,2,1]。
        :param max_class_rank: 最大班级次，整数。默认60。
        :param rev:翻转列的顺序。
        :return: df表。班级分析报表。

        """

        # 定义偏函数，设置阈值列表，默认为[0,200,400]。cumu:0为不累计，1为累计。
        funs_cls_0 = functools.partial(funs_fd, thresh=thresh,cumu=0,mode=1,inf=1,rev=rev)
        funs_cls_1 = functools.partial(funs_fd, thresh=thresh,cumu=-1,mode=1,inf=1,rev=rev)

        # 获取班级名次表["班级"，"级次"]
        df_cls = self.get_df(columns=["班级" ,"总分"],max_class_rank=max_class_rank)
        # 用班级分组，用总分应用名次段函数,count函数用于统计数据总个数。
        df_cls = df_cls.groupby("班级")["总分"].agg(funs_cls_0()+funs_cls_1())

        if thresh_score !=0:
            # 设置偏函数，为点积函数设置b值为默认值.
            fun_jf = functools.partial(np.dot, b=thresh_score)  # 设置偏函数,预先设置qz的值为JF_FSD_SDB.
            # 插入参评人数列
            df_cls.insert(loc=0, column="参评人数", value=df_cls.iloc[:, 0:(len(thresh))].apply(sum, axis=1))
            # 添加积分列
            df_cls.insert(loc=1, column="积分",value=df_cls.iloc[:, 1:(len(thresh_score) + 1)].apply(fun_jf, axis=1))
            # 添加排名列.
            df_cls.insert(loc=2, column="排名",value=df_cls["积分"].rank(axis=0, ascending=False, method="min"))

        return df_cls

def fun_sdb(sr, s_rank, t_rank):
    """
    返回满足 x <= s_rank 且 y <= t_rank 的元组个数。
    :param sr: pd.Series, 数据列表，[(x1, y1),(x2, y2),(x3, y3),(x4, y4)...],每个元素是一个元组 (xn, yn)。
    :param s_rank: int, 学科排名。
    :param t_rank: int, 总分排名。
    :return: int, 满足条件的元组个数。
    """
    # 类型检查
    if not isinstance(sr, pd.Series):
        raise ValueError("sr 必须是 pd.Series 类型")
    if not isinstance(s_rank, int) or not isinstance(t_rank, int):
        raise ValueError("s_rank 和 t_rank 必须是整数")
    if not all(isinstance(x, tuple) and len(x) == 2 for x in sr):
        raise ValueError("sr 中的每个元素必须是长度为 2 的元组")

    # 将 Series 转换为 NumPy 数组
    arr = np.array(sr.tolist())
    # 计算满足条件的元组个数
    res = np.sum((arr[:, 0] <= s_rank) & (arr[:, 1] <= t_rank))
    return res
def funs_fd(thresh, cumu=0, mode=0, inf=1,rev=0):
    """
    返回一组函数，每个函数统计落在指定区间的数据个数。
    :param thresh: 成绩阈值的列表，如 [0, 60, 80, 100, 120]
    :param cumu: cumulative,整数类型,0或1,数据是否累计，默认不累计.
                 1: 累计统计（小于等于或小于某个值）
                 -1: 反向累计统计（大于等于或大于某个值）
                 0: 非累计统计（落在某个区间内）
    :param mode: 整数类型,0或1.=0时,前闭后开,=1时,前开后闭.
    :param inf: 整数类型.0或1.=0,不添加正无穷,=1,添加正无穷
    :param rev: 整数类型，0或1。=0，不翻转函数列表，=1，翻转函数列表。
    :return: 一组函数，每个函数统计落在某个区间的数据个数。
             每个函数的形式为 FS_xx-xx(arr)，其中 arr 为待统计的列表数据。
    """
    # 确保 thresh 是一个列表或元组
    if not isinstance(thresh, (list, tuple)):
        raise ValueError("thresh 必须为列表或元组")

    # 在阈值列表末尾添加正无穷，以处理最后一个区间
    thresh = list(thresh) + [float('inf')] if inf == 1 else list(thresh)

    # 确保 thresh 至少有两个元素
    if len(thresh) < 2:
        raise ValueError("thresh 列表至少需要两个元素")

    def make_counter(lower, upper):
        def counter(arr):
            arr = np.array(arr)
            if cumu == 1:
                # 累计统计：计算在 mode == 0 [-∞, upper) 或 mode == 1 (-∞, upper] 内的元素个数
                if mode == 0:
                    return np.sum(arr < upper)
                else:
                    return np.sum(arr <= upper)
            elif cumu == -1:
                # 反向累计统计：计算在 mode == 0 (lower, ∞) 或 mode == 1 [lower, ∞) 内的元素个数
                if mode == 0:
                    return np.sum(arr > lower)
                else:
                    return np.sum(arr >= lower)
            else:
                # 非累计统计：计算在 mode == 0 [lower, upper) 或 mode == 1 (lower, upper] 内的元素个数
                if mode == 0:
                    return np.sum((arr >= lower) & (arr < upper))
                else:
                    return np.sum((arr > lower) & (arr <= upper))

        # 根据 mode 确定区间的开闭符号
        # left_bracket = "[" if mode == 0 else "("
        # right_bracket = ")" if mode == 0 else "]"
        if mode == 0:
            left_bracket,right_bracket  = "[",")"
        else:
            left_bracket, right_bracket = "(", "]"


        # 根据 cumu 的值确定统计类型
        if cumu == 1:
            stat_type = "Cum"  # 累计统计
            lower=''
        elif cumu == -1:
            stat_type = "CumRev"  # 反向累计统计
            upper =''
        else:
            stat_type = "Cnt"  # 非累计统计

        # 设置函数名
        counter.__name__ = f"Fd{left_bracket}{lower}-{upper}{right_bracket}{stat_type}"
        return counter

    # 创建所有区间的统计函数
    funcs = [make_counter(thresh[i], thresh[i + 1]) for i in range(len(thresh) - 1)]
    funcs = funcs if rev == 0 else funcs[::-1]
    return funcs
def funs_mcd_tup(thresh, max_total_rank):
    """
    返回一个函数组，每个函数统计落在指定分数区间内的数据个数（前开后闭）。
    :param thresh: 列表类型，学科名次的阈值列表。必须是升序排列。如:[0,60, 80, 100]
    :param max_total_rank: 整数或浮点数，总分名次。
    :return: 一个函数列表。每个函数统计落在某个名次段中的数据个数,且要求小于总名次。
    """
    if not isinstance(thresh, (list, tuple)) or not all(isinstance(x, (int, float)) for x in thresh):
        raise ValueError("thresh 必须为升序排列的整数或浮点数列表")
    if not isinstance(max_total_rank, (int, float)):
        raise ValueError("max_total_rank 必须为整数或浮点数")
    if any(thresh[i] >= thresh[i + 1] for i in range(len(thresh) - 1)):
        raise ValueError("thresh 必须是升序排列的")

    def make_counter(lower, upper):
        def counter(arr):
            return sum(1 for x, y in arr if lower < x <= upper and y <= max_total_rank)
        counter.__name__ = f'MC（{lower}-{upper}］T≤{max_total_rank}'
        # print(counter.__name__)
        return counter

    def last_counter(arr):
        return sum(1 for _, y in arr if y > max_total_rank)

    last_counter.__name__ = f'MC T>{max_total_rank}'

    # 在阈值列表末尾添加正无穷
    thresh = list(thresh) + [float('inf')]
    # 生成计数器函数组
    counters = [make_counter(thresh[i], thresh[i + 1]) for i in range(len(thresh) - 1)] + [last_counter]
    return counters
def funs_lv(thres):
    """
    返回一个不同名的函数组，如计算:及格人数,及格率,优秀人数,优秀率。
    :param thres: 列表类型。是一个成绩阈值的列表，如: [60, 80, 100]。用于计算及格率,优秀率
    :return: 一个不同名的函数组，每个函数统计及格率,优秀率。
    """
    if not isinstance(thres, (list, tuple)):
        raise ValueError("thres 必须为列表或元组")

    def make_func(__lower, __upper, ratio=0):
        """
        创建一个用于统计人数或比率的函数
        :param __lower: 区间的下界
        :param __upper: 区间的上界
        :param ratio: 0 表示统计人数，1 表示统计比率
        :return: 一个统计函数
        """

        def __func(__scores):
            arr = np.array(__scores)
            count = np.sum((arr >= __lower) & (arr <= __upper))
            if ratio == 1:
                return count / len(arr) if len(arr) > 0 else 0
            else:
                return count

        __func.__name__ = f'Lv[{__lower}-{__upper}]{"_ratio" if ratio == 1 else "_count"}'
        return __func

    # 创建所有区间的统计函数
    funcs = []
    for i in range(len(thres) - 1):
        __lower = thres[i]
        __upper = thres[-1]
        funcs.append(make_func(__lower, __upper, ratio=0))
        funcs.append(make_func(__lower, __upper, ratio=1))
    return funcs
def trim_wb(wb, sheet_name):
    """
    只保留指定名称的工作表，并删除工作簿中的其他所有工作表。
    :param wb: openpyxl工作簿对象
    :param sheet_name: 要保留的工作表名称
    :return: 修改后的工作簿对象以及保留的工作表对象
    :raises ValueError: 如果指定的工作表名称不存在于工作簿中
    """
    # 检查指定的工作表是否存在
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"工作表 '{sheet_name}' 不存在于工作簿中")

    # 删除其他工作表
    for sheet in list(wb.worksheets):
        if sheet.title != sheet_name:
            wb.remove(sheet)

    # 返回工作簿和保留的工作表对象
    return wb, wb[sheet_name]
def sr_to_ws(ws, cells, data):
    """
    把列表或pandas的Series数据填充到openpyxl创建的ws表的指定单元格中。

    :param ws: 由openpyxl创建的Worksheet对象。
    :param cells: 一个坐标列表，如[(2, 2), (2, 4), (3, 3), (3, 11)]，表示要填充的单元格位置。
    :param data: 一个Python列表或Pandas Series对象，包含要填充的值。
    :return: 无返回值

    注意：确保data的长度与cells的长度相匹配，否则将引发异常。
    """
    # 检查ws是否为Worksheet对象
    if not hasattr(ws, 'cell'):
        raise ValueError("ws必须是一个openpyxl的Worksheet对象")

    # 检查data的类型
    if isinstance(data, pd.Series):
        # 处理Pandas Series中的值，将NaN转换为None（或其他占位符）
        values = data.tolist()
    elif isinstance(data, list):
        values = data
    else:
        raise ValueError("data必须是一个Pandas Series对象或Python列表")

    # 检查cells和data的长度是否相等
    if len(cells) != len(values):
        raise ValueError("cells和data的长度必须相等")

    # 遍历cells和values，将值填充到对应的单元格中
    for (row, col), val in zip(cells, values):
        # 填充单元格
        ws.cell(row=row, column=col, value=val)
def dfs_to_ws(ws, row, col, dfs, rg=10, cg=0, hd=False):
    """
    将一个或多个Pandas DataFrame逐一注入到一个由openpyxl创建的工作表中。

    :param ws: 由openpyxl创建的工作表。
    :param row: 第一个DataFrame的行坐标（起始行）。
    :param col: 第一个DataFrame的列坐标（起始列）。
    :param dfs: 单个DataFrame或其所组成的列表。
    :param rg: DataFrame之间的行间距。
    :param cg: DataFrame之间的列间距（通常保持为0）。
    :param hd: 是否包含DataFrame的列头，默认为False。
    :return: None. 只有注入操作，返回值为空。
    """
    # 检查ws是否为Worksheet对象
    if not hasattr(ws, 'cell'):
        raise ValueError("ws必须是一个openpyxl的Worksheet对象")

    # 将单个DataFrame转换为列表
    if isinstance(dfs, pd.DataFrame):
        dfs = [dfs]

    # 检查dfs是否为DataFrame列表
    if not all(isinstance(df, pd.DataFrame) for df in dfs):
        raise ValueError("dfs必须是一个Pandas DataFrame对象或包含Pandas DataFrame对象的列表")

    # 遍历每个DataFrame
    for df in dfs:
        # 遍历DataFrame的每一行
        for r, row_data in enumerate(dataframe_to_rows(df, index=False, header=hd), start=row):
            # 将每行数据写入到工作表的对应位置
            for c, value in enumerate(row_data, start=col):
                # 将NaN值转换为None
                if pd.isnull(value):
                    value = None
                ws.cell(row=r, column=c, value=value)

        # 更新行位置为下一个DataFrame的起始位置
        row += rg
        col += cg
def df_split(df, level=0):
    """
    将具有多级列索引的DataFrame按照第0级列索引拆分成多个单独的DataFrame。

    参数:
        df (pd.DataFrame): 输入的DataFrame，应该具有多级列索引。
        level (int): 拆分DataFrame的层级，默认为0（虽然这里只支持0，但保留参数以备将来扩展）。
    返回:
        List[pd.DataFrame]: 包含多个DataFrame的列表，每个DataFrame包含具有相同第0级列索引的列。
    异常:
        ValueError: 如果输入的DataFrame没有多级列索引或指定的层级无效。
    """
    # 确保DataFrame具有多级列索引
    if not isinstance(df.columns, pd.MultiIndex):
        raise ValueError("输入的DataFrame必须具有多级列索引")
    # 验证指定的层级是否有效
    if level < 0 or level >= df.columns.nlevels:
        raise ValueError(f"指定的层级 {level} 超出了DataFrame的列索引范围")

    # 如果指定的层级不是0，交换列索引的层级
    if level != 0:
        df = df.swaplevel(level, 0, axis=1)
    # 获取第0级的所有唯一标签
    unique_labels = df.columns.get_level_values(0).unique()

    # 初始化一个空列表来存储拆分后的DataFrame
    split_dfs = []
    # 遍历每个唯一标签
    for label in unique_labels:
        # 选择所有在第0级具有该标签的列
        selected_columns = df.columns[df.columns.get_level_values(0) == label]
        # 使用选定的列创建一个新的DataFrame
        temp_df = df[selected_columns]
        # 将新的DataFrame添加到列表中
        split_dfs.append(temp_df)
    return split_dfs
def df_groupby(df: pd.DataFrame, col: str) -> dict[str, pd.DataFrame]:
    """
    按指定列将 DataFrame 分割成多个子 DataFrame，并将它们存储在一个字典中。

    :param df: DataFrame 对象
    :param col: 用于分组的列标签
    :return: 一个字典，键为组名，值为对应的子 DataFrame
    """
    if not isinstance(df, pd.DataFrame):
        raise ValueError("df 必须是 DataFrame 对象")
    if col not in df.columns:
        raise ValueError(f"列标签 '{col}' 不在 DataFrame 中")

    return {name: group.reset_index(drop=True) for name, group in df.groupby(col)}
def wb_to_bytesIO(wb):
    """
    把一个 openpyxl 生成的 workbook 对象封装为一个二进制的 BytesIO 对象。

    :param wb: 由 openpyxl 生成的 workbook 对象
    :return: 一个二进制的 BytesIO 对象
    """
    if not isinstance(wb, openpyxl.Workbook):
        raise ValueError("wb 必须是 openpyxl 生成的 Workbook 对象")

    bio_file = BytesIO()
    wb.save(bio_file)
    bio_file.seek(0)  # 确保指针位于文件的开头
    return bio_file
def dfs_to_zip(
        dfs_dic: Dict[str, pd.DataFrame],
        compression: int = zipfile.ZIP_DEFLATED,
        format: str = 'excel',
        empty_message: str = '空值') -> BytesIO:
    """
    将多个DataFrame保存到内存中的ZIP文件

    Parameters:
    -----------
    dfs_dic : dict
        包含DataFrame的字典，键为文件名（不含扩展名）
    compression : int, optional
        压缩方法，默认为zipfile.ZIP_DEFLATED
    format : str, optional
        输出格式，支持'excel'（默认）或'csv'
    empty_message : str, optional
        空DataFrame时显示的消息，默认为'该班级数据为空'

    Returns:
    --------
    BytesIO
        包含ZIP文件内容的字节缓冲区

    Raises:
    -------
    ValueError
        如果指定的格式不被支持
    """
    # 验证格式参数
    if format not in ('excel', 'csv'):
        raise ValueError(f"不支持的格式: {format}. 支持 'excel' 或 'csv'")

    # 创建内存中的ZIP文件
    bio_zip = BytesIO()

    try:
        with zipfile.ZipFile(bio_zip, 'w', compression) as zipf:
            for name, df in dfs_dic.items():
                # 检查DataFrame是否为空
                if df.empty:
                    df = pd.DataFrame({'提示': [empty_message]})

                # 安全处理文件名
                safe_name = _sanitize_filename(str(name) if name is not None else 'data')

                # 根据格式处理数据
                if format == 'excel':
                    file_data = _df_to_excel(df, safe_name)
                    file_ext = 'xlsx'
                else:  # csv
                    file_data = _df_to_csv(df)
                    file_ext = 'csv'

                # 将文件数据写入ZIP
                zipf.writestr(f'{safe_name}.{file_ext}', file_data)

    except Exception as e:
        # 重新抛出异常，但先确保缓冲区被重置
        bio_zip.seek(0)
        bio_zip.truncate(0)
        raise e

    # 将指针重置到缓冲区开头
    bio_zip.seek(0)
    return bio_zip

def _sanitize_filename(name: str, max_length: int = 30) -> str:
    """
    安全处理文件名，移除非法字符并限制长度

    Parameters:
    -----------
    name : str
        原始文件名
    max_length : int, optional
        最大长度限制，默认为30

    Returns:
    --------
    str
        处理后的安全文件名
    """
    # 移除非ASCII字符和非法文件名字符
    safe_name = re.sub(r'[^\w\s-]', '', name).strip()

    # 替换空格为下划线
    safe_name = re.sub(r'[-\s]+', '_', safe_name)

    # 限制长度
    if len(safe_name) > max_length:
        safe_name = safe_name[:max_length]

    # 如果为空则使用默认名称
    if not safe_name:
        safe_name = 'data'

    return safe_name
def _df_to_excel(df: pd.DataFrame, sheet_name: str) -> bytes:
    """
    将DataFrame转换为Excel字节数据

    Parameters:
    -----------
    df : pd.DataFrame
        要转换的DataFrame
    sheet_name : str
        Excel工作表名称

    Returns:
    --------
    bytes
        Excel文件的字节数据
    """
    excel_buffer = BytesIO()

    try:
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)  # Excel限制工作表名31字符
        return excel_buffer.getvalue()
    finally:
        excel_buffer.close()
def _df_to_csv(df: pd.DataFrame) -> bytes:
    """
    将DataFrame转换为CSV字节数据

    Parameters:
    -----------
    df : pd.DataFrame
        要转换的DataFrame

    Returns:
    --------
    bytes
        CSV文件的字节数据（UTF-8编码）
    """
    csv_buffer = BytesIO()

    try:
        # 使用UTF-8编码确保中文正确显示
        df.to_csv(csv_buffer, index=False, encoding='utf-8-sig')
        return csv_buffer.getvalue()
    finally:
        csv_buffer.close()


def df_split_column(df: pd.DataFrame, col: List[str]) -> Dict[str, pd.DataFrame]:
    """
    将一个DataFrame拆分为多个DataFrame对象，每个对象包含一个指定列和所有其他列

    Parameters:
    -----------
    df : pd.DataFrame
        要拆分的原始DataFrame
    col : List[str]
        要拆分的列名列表

    Returns:
    --------
    Dict[str, pd.DataFrame]
        字典，键为拆分列名，值为对应的DataFrame对象

    Raises:
    -------
    ValueError
        如果指定的列不在DataFrame中
    """
    # 检查所有指定的列是否存在于DataFrame中
    missing_cols = [c for c in col if c not in df.columns]
    if missing_cols:
        raise ValueError(f"以下列不存在于DataFrame中: {missing_cols}")

    # 获取所有不在col列表中的列
    other_cols = [c for c in df.columns if c not in col]

    # 创建结果字典
    result = {}

    # 为每个指定列创建一个新的DataFrame
    for c in col:
        # 选择当前列和所有其他列
        selected_cols = other_cols + [c]
        result[c] = df[selected_cols].copy()

    return result




