import functools
import math
import warnings
from io import BytesIO
import numpy as np
import openpyxl
import pandas as pd
import streamlit
from openpyxl.utils.dataframe import dataframe_to_rows
import zipfile
import re
from typing import Dict, Union, List, Any, Callable, Optional, Tuple, Literal
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


'''
# 2024.07.08更新：对多个函数进行优化。
# 2024.11.04更新：对整个模块进行重构，增强Andf功能，优化相关功能函数。
# 2024.11.24更新：完成als重构，重命名为al模块。
# 2025.02.01更新：添加了get_cls_score()方法，实现了班级分析-分数段功能；给funs_fd()添加rev参数。
# 2025.09.15更新：修改添加了多个函数，是一次巨大的提升。
'''

# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■      Andf类及其方法      ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 创建Andf类,以生成多种报表:两率一平报表,学科分数段报表,学科双达标报表,班级名次段报表
class Andf:

    # 标准学科名列表：self.__sbj，           包括成绩表没有的学科名。
    # 标准学科名字典：self.__sbj_dic，       标准学科名对应的字典。
    # 已用学科名列表：self.__sbj_lst，       所有已经使用的学科名列表。
    # 总分班次级次成绩全表：self.__df         包括总分、班次、级次的成绩表。
    def __init__(self, df: pd.DataFrame) -> None:

        self.__df = df
        # 声明标准学科名列表
        self.__sbj: List[str] = ["语文", "数学", "英语", "物理", "化学", "生物", "政治", "历史", "地理"]
        # 确定df表对应的[学科名]列表:__sbj_lst.
        self.__sbj_lst: List[str] = list(set(self.__sbj) & set(self.__df.columns))
        # 确定df表对应的{学科名：序号}字典：__sbj_dic。
        self.__sbj_dic: Dict[str, int] = {val: idx + 1 for idx, val in enumerate(self.__sbj)}
        # 对学科列表进行排序，排序规则为：字典中key对应的值。
        self.__sbj_lst.sort(key=lambda x: self.__sbj_dic[x])

        # 确保学科成绩列都是数值类型，避免在计算总分时出现类型错误
        for subject in self.__sbj_lst:
            self.__df[subject] = pd.to_numeric(self.__df[subject], errors='coerce')
        
        # 增加总分列、班次、级次列，创建基础df对象
        self.__df["总分"] = df.loc[:, self.__sbj_lst].sum(axis=1, min_count=1)
        # 增加班次列
        self.__df["班次"] = df.groupby("班级")["总分"].rank(ascending=False, method="min")
        # 增加级次列
        self.__df["级次"] = df["总分"].rank(axis=0, ascending=False, method="min")

    # ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■   基本信息表（学科表，学科字典，全信息表） ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
    # ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
    # 获取df表存在的标准学科名列表，己正确排序。
    def get_sbj_lst(self) -> List[str]:
        """
        返回值：学科列表。标准学科名称对应的列表。
        """
        # 对学科列表进行排序，排序规则为：字典中key对应的值。
        return self.__sbj_lst

    # ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
    # 获取学科名_排序的字典(学科名：序号)
    def get_sbj_dic(self) -> Dict[str, int]:
        """
        返回值：学科名：排序的字典。标准学科名称对应的字典("语文":1，"数学":2，"英语":3，"物理":4，"化学":5，"生物":6，"政治":7，"历史":8，"地理":9)
        """
        return dict((key, self.__sbj_dic[key]) for key in self.__sbj_lst)

    # ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
    # 获取全部信息的df表
    def get_all(self) -> pd.DataFrame:
        """
        返回导入后，添加总分、班次、校次列后的DF表。
        :return: 添加总分、班次、校次列后的DF表。
        """
        return self.__df
    # ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■   基础报表（部分信息表，名次表，学科双达标表）  ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
    # ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
    # 按列名列表与最大班次获取部分df表
    def get_df(self,
               columns: Optional[List[str]] = None,   # 例：["班级", "学号", "姓名",'数学']
               max_class_rank: int = 60               # 获取班级分析报表中时，最大计算人数。
               ) -> pd.DataFrame:
        """
        返回值：df表。
        :param columns: 列名，列表。
        :param max_class_rank: 班级数，整数。默认40。
        :return: df表。
        """
        if columns is None:
            columns = ["班级", "学号", "姓名"]
        df = self.__df[self.__df["班次"] <= max_class_rank]
        return df.loc[:, columns]

    # ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
    # 获取名次表（学科名次，总分名次）
    def get_mc(self,
               max_class_rank: int = None,     # 获取班级分析报表中时，最大计算人数。
               combine_ranks: int = 1          # 是否合并学科名次，总分名次为元组。默认1：合并.
               ) -> pd.DataFrame:
        """
        返回班次满足小于max_class_rank的df表.当combine_ranks=1合并班次与总分名次为一个元组
        :param max_class_rank: 班级名次，整数。默认40,取班次<=40的数据。
        :param combine_ranks: 整数1或0。0,各科独立排名。默认1,学科排名与总分排名组合为元组.
        :return: df表。班次满足小于max_class_rank的df表.当combine_ranks=1合并班次与总分名次为一个元组
        """
        if max_class_rank != None:
            df = self.__df[self.__df["班次"] <= max_class_rank]
        else:
            df= self.__df.copy()
        df_mc = df_rank_cols(df, self.__sbj_lst+["总分"], method='min', ascending=False)

        if combine_ranks == 1:
            df_mc = df_pair_cols(df_mc, self.__sbj_lst, '总分')

        return df_mc


    # ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
    # 获取学科双达标报表
    def get_db(self,
               max_subject_rank: int = 40,
               max_total_rank: int = 40
               ) -> pd.DataFrame:
        """
        返回班级分析报表.
        :param max_subject_rank: 科目最大名次，整数。默认40.
        :param max_total_rank: 总分最大名次，整数。默认40.
        :return: 班级分析报表.
        """
        # 获取班级名次（学科名次，总分名次）表。
        df_mc = self.get_mc(max_class_rank=None, combine_ranks=1)
        df_mc = df_mc[["班级"]+self.__sbj_lst]

        # 获取学科双达标函数列表
        func = functools.partial(count_dual_cond, a=max_subject_rank, b=max_total_rank)
        df_db = df_mc.groupby("班级").agg(func)

        return df_db

    # ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■   4大核心报表（多阈值报表）■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
    # ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
    # 生成学科多阈值分数段报表.
    def get_fsd(self,
                dic_thresh_sbj: Dict[Tuple, List[str]],            # 阈值列表.如:[0,36,72,96,120]
                thresh_score: List[Union[int, float]] = None,      # 积分列表.如:[10, 9, 2, 1, 0]
                max_class_rank: int = 60,       # 班级最大名次，整数。默认40.
                add_rank_cols = 1                 # 是否加入求各，点积、排名列.
                ) -> Dict[Any, pd.DataFrame]:
        """
        生成分数段报表.
        :param dic_val_sbj: 阈值列表.如:[0,36,72,96,120]
        :param thresh_score: 学科名次段阈值列表对应积分.如:[10, 9, 2, 1, 0]
        :param max_class_rank: 最大班次,整数.即班级最大参评人数
        :return: 分数段报表.
        """
        # 生成字典{学科:（72，96，120）}，并排序
        dic = dict_rev_sort(dic=dic_thresh_sbj, sort_order=self.__sbj_lst)

        # 定义{学科:[分数段函数组]}字典
        dic_FAN = {key: functools.partial(make_bin_counters, thresh=dic[key])() for key in dic.keys()}

        # 获取df表,要求最大班次小于max_class_rank,列索引中有班级与学科名.
        df = self.get_df(columns=["班级"] + self.get_sbj_lst(), max_class_rank=max_class_rank)

        # 按班级分组后，执行分数段函数组。生成一个报表.
        df_fsd = df.groupby("班级").agg(dic_FAN)

        # 依据第0列索引，分割数据为多个df表。
        df_fsd = df_split_levels(df_fsd)

        for key, dff in df_fsd.items():
            # 添加各分排名三列
            if thresh_score != None and  add_rank_cols == 1 :
                dff = df_add_rank(dff, lst=thresh_score, sum_col_name="积分", dot_col_name="点积", rank_col_name="点积排名")
            # 添加行索引名称
            dff.index.name = key
            # 更新字典中的DataFrame
            df_fsd[key] = dff


        return df_fsd

    # 生成学科多阈值双达标报表.
    def get_sdb(self,
                thresh: List[int],                              # 阈值列表.如:[0,200,260,300]
                thresh_score: List[Union[int, float]] = [10, 9, 2, 1, 0],     # 积分列表.如:[10, 9, 2, 1, 0]
                max_class_rank: int = None,                       # 班级最大名次，整数。默认60.
                max_total_rank: int = 200                       # 全校最大名次，整数。默认200.
                ) -> Dict[Any, pd.DataFrame]:
        """
        生成双达标报表.
        :param thresh: 阈值列表.如:[0,200,260,300]
        :param thresh_score: 学科名次段阈值列表对应积分.如:[10, 9, 2, 1, 0]
        :param max_class_rank: 最大班次,整数.即班级最大参评人数
        :param max_total_rank: 最大校次.即最大校次,如:260
        :return: 双达标报表.
        """

        # 获取班级名次（学科名次，总分名次）表。
        df_mc = self.get_mc(max_class_rank=max_class_rank, combine_ranks=1)

        # 获取学科双达标函数列表
        mcd_funs = functools.partial(make_dual_cond_counters, thresh=thresh, sec_thresh=max_total_rank)

        # 创建字典，添加键值对为{学科名：单达标函数},以计算单达标人数.
        dic_mcd_funs = {str(i): mcd_funs() for i in self.__sbj_lst}

        # 使用聚合函数，计算学科双达标人数，并返回df表。
        df_SDB = df_mc.groupby("班级").agg(dic_mcd_funs)

        # 依据第0列索引，分割数据为多个df表。
        dfs_xk = df_split_levels(df_SDB)

        # 添加各科积分列
        if thresh_score != 0:
            for key, dff in dfs_xk.items():
                # 添加各分排名三列
                dff = df_add_rank(dff, lst=thresh_score, sum_col_name="积分", dot_col_name="点积", rank_col_name="排名")
                # 添加行索引名称
                dff.index.name = key
                # 更新字典中的DataFrame
                dfs_xk[key] = dff
        return dfs_xk

    # 生成两率一平报表.
    def get_lv(self,
               dic_total_sbj: Dict[int, List[str]],     #
               thresh: List[float] = [0.6, 0.8],        # 阈值列表.如:[0.6,0.8]
               max_class_rank: int = 40,                # 班级最大名次，整数。默认40.
               include_count_valid: int = 0,             # 添加统计有效人数列。默认不统计：0。
               add_rank_cols= None,
               ) -> Dict[Any, pd.DataFrame]:
        """
        生成两率一平报表.
        :param dic_total_sbj: {学科：总分}字典。如:{120: ["语文", "数学", "英语"], 70: ["物理", "政治"], 50: ["化学", "生物", "历史", "地理"]}
        :param calcu: 0或1，默认1。0，添加。1，添加分析列。
        :param thresh: 及格率优秀率或其它比率的阈值,例[0.6,0.8]
        :param max_class_rank: 最大班次次，整数。默认40,取班次<=40的数据。
        :return: df表。
        """

        # 使用字典推导式转换为{学科:总分}字典
        dic = dict_rev_sort(dic=dic_total_sbj, sort_order=self.__sbj_lst)

        # 定义{学科:[两率一平函数组]}字典
        # 明确设置 include_mean=True
        dic_FAN = {key: functools.partial(make_rate_counters,
                                          thresh=np.array(thresh + [1]) * val,
                                          include_count_valid=include_count_valid,
                                          include_mean=True)() for key, val in dic.items()}

        # 获取需进行两率一平的df表,要求最大班次小于max_class_rank,列索引中人班级与学科名.
        df = self.get_df(columns=["班级"] + self.get_sbj_lst(), max_class_rank=max_class_rank)

        # 按班级分组后，执行两率一平计算。生成一个报表.
        df_lv = df.groupby("班级").agg(dic_FAN)

        # 依据第0列索引，分割数据为多个df表。
        dfs_lv = df_split_levels(df_lv)

        # 加入排名列
        if add_rank_cols !=None:
            for key, dff in dfs_lv.items():
                dfs_lv[key] = df_add_cols_rank(dff, columns_to_rank=add_rank_cols)

        return dfs_lv

    # 生成班级分析报表-各次段统计
    def get_cls(self,
                thresh: List[int],
                thresh_score: List[Union[int, float]] = None,
                max_class_rank: int = 60,
                cumu: int = 0,
                mode: int = 1
                ) -> pd.DataFrame:
        """
        获取班级分析报表-各次段统计。
        :param thresh: 名次段阈值列表，列表。默认[0,100,200,300]。
        :param thresh_score: 名次段积分列表，列表。默认[4,3,2,1]。
        :param max_class_rank: 最大班级次，整数。默认60。
        :param cumu:  累计开关：0为不累计，1为累计。
        :param mode: 区间模式：0 为前闭后开，1为前开后闭。
        :return: df表。班级分析报表。
        """

        # 定义偏函数，设置阈值列表，若为0，则默认为[0,200,400]。cumu:0为不累计，1为累计。
        funs_cls = functools.partial(make_bin_counters, thresh=thresh, cumu=cumu, mode=mode)

        # 获取班级名次表["班级"，"级次"]
        df_cls = self.get_df(columns=["班级", "级次"], max_class_rank=max_class_rank)
        # 按班级分组后，执行班级名次计算。生成一个报表.
        df_cls = df_cls.groupby("班级")["级次"].agg(funs_cls())

        if thresh_score is not None:
            # 添加各分排名三列
            df_cls = df_add_rank(df_cls, lst=thresh_score, sum_col_name="总人数", dot_col_name="点积", rank_col_name="排名")
        return df_cls

    # ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■   组合报表（基本报表+核心报表） ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
    # ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
    def get_db_fsd(self,
                   dic_thresh_sbj: Dict[Tuple, List[str]],  # 阈值列表.如:[0,36,72,96,120]
                   thresh_score: List[Union[int, float]] = None,  # 积分列表.如:[10, 9, 2, 1, 0]
                   max_subject_rank: int = 40,
                   max_total_rank: int = 40
                   )->  Dict[Any, pd.DataFrame]:

        # 获取学科达标表
        df_db = self.get_db( max_subject_rank=max_subject_rank, max_total_rank=max_total_rank)
        fsd_df = self.get_fsd(dic_thresh_sbj=dic_thresh_sbj,thresh_score=thresh_score,add_rank_cols=0)
        for key,dff in fsd_df.items():
            dff.insert(0, '双达标<='+ str(max_subject_rank), df_db[key])
            dff = df_add_rank(dff, lst=thresh_score, sum_col_name="积分", dot_col_name="点积", rank_col_name="排名")
            fsd_df[key] = dff
        return fsd_df


# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■   四大分析函数（组）单元   ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 统计 Series 中同时满足两个条件的元素个数
def count_dual_cond(
        sr: pd.Series,
        a: Union[int, float],
        b: Union[int, float],
        op1: str = 'le',
        op2: str = 'le') -> int:
    """
    2025.09.18 修改.来源于fun_sdb.
    📊 统计 Series 中同时满足两个条件的元素个数。
    每个元素为 (x, y) 二元组，统计满足 x ◎1 a 且 y ◎2 b 的个数。

    📌 支持操作符：
        'le' → <= （默认，越小越好，如排名）
        'ge' → >= （越大越好，如分数）
        'lt' → <  （严格小于）
        'gt' → >  （严格大于）

    📝 参数：
        sr (pd.Series): 元素为 (x, y) 二元组的 Series
        a (Real): 第一个条件的阈值
        b (Real): 第二个条件的阈值
        op1 (str): x 的比较操作符，默认 'le'
        op2 (str): y 的比较操作符，默认 'le'

    📤 返回：
        int: 同时满足两个条件的元素个数

    🧪 示例：
        sr = pd.Series([(1,11), (2,22), (3,33), (4,44)])
        count_dual_cond(sr, 3, 33, op1='le', op2='le')  # 返回 3 → (1,11), (2,22), (3,33)
        count_dual_cond(sr, 3, 33, op1='ge', op2='ge')  # 返回 2 → (3,33), (4,44)

    🚨 异常：
        ValueError: 当参数不符合要求时
        TypeError: 当参数类型不正确时
    """
    # === 🚦 阶段1：基础类型校验 ===
    # 1. 验证 sr 类型
    if not isinstance(sr, pd.Series):
        raise TypeError("sr 必须是 pd.Series 类型")

    # 2. 验证 a 和 b 类型
    if not isinstance(a, (int, float)):
        raise TypeError("a 必须是数值类型（int 或 float）")
    if not isinstance(b, (int, float)):
        raise TypeError("b 必须是数值类型（int 或 float）")

    # 3. 验证 a 和 b 值
    if np.isnan(a) or np.isnan(b):
        raise ValueError("a 和 b 不能为 NaN")
    if np.isinf(a) or np.isinf(b):
        raise ValueError("a 和 b 不能为无穷大")

    # 4. 验证 op1 和 op2 类型
    if not isinstance(op1, str):
        raise TypeError("op1 必须是字符串")
    if not isinstance(op2, str):
        raise TypeError("op2 必须是字符串")

    # === 🧮 阶段2：定义操作符映射 + 校验操作符合法性 ===
    op_map = {
        'le': lambda x, t: x <= t,  # 小于等于
        'ge': lambda x, t: x >= t,  # 大于等于
        'lt': lambda x, t: x < t,  # 小于
        'gt': lambda x, t: x > t  # 大于
    }

    if op1 not in op_map:
        raise ValueError(f"op1 必须是 'le', 'ge', 'lt', 'gt' 之一，当前值: {op1}")
    if op2 not in op_map:
        raise ValueError(f"op2 必须是 'le', 'ge', 'lt', 'gt' 之一，当前值: {op2}")

    # === 📋 阶段3：数据内容校验 ===
    if len(sr) == 0:
        return 0

    # 过滤掉包含 NaN 或无穷大的元素
    valid_items = []
    for i, item in enumerate(sr):
        # 检查是否为元组
        if not isinstance(item, tuple):
            raise ValueError(f"Series 中每个元素必须是元组，但索引 {i} 的元素是 {type(item).__name__}")

        # 检查元组长度
        if len(item) != 2:
            raise ValueError(f"Series 中每个元素必须是长度为2的元组，但索引 {i} 的元素长度为 {len(item)}")

        # 检查元组元素类型
        x, y = item
        if not isinstance(x, (int, float)):
            raise ValueError(f"Series 中每个元素的第一个值必须是数字，但索引 {i} 的第一个值是 {type(x).__name__}")
        if not isinstance(y, (int, float)):
            raise ValueError(f"Series 中每个元素的第二个值必须是数字，但索引 {i} 的第二个值是 {type(y).__name__}")

        # 检查元组元素值，跳过包含 NaN 或无穷大的元素
        if np.isnan(x) or np.isnan(y) or np.isinf(x) or np.isinf(y):
            continue  # 跳过包含无效值的元素

        valid_items.append(item)

    # 如果没有有效数据，返回 0
    if not valid_items:
        return 0

    # === 🚀 阶段4：执行计算（NumPy 向量化，高性能）===
    try:
        arr = np.array(valid_items)  # 转为二维数组 (n, 2)
        mask = op_map[op1](arr[:, 0], a) & op_map[op2](arr[:, 1], b)  # 构建布尔掩码
        return int(np.sum(mask))  # 统计 True 的个数并返回 Python int
    except Exception as e:
        raise RuntimeError(f"计算过程中发生错误: {str(e)}") from e

# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 生成一组按数值区间（bin）进行计数的函数，支持滑动区间与累计区间、前闭后开与前开后闭模式。
def make_bin_counters(
        thresh: Union[List[Union[int, float]], Tuple[Union[int, float], ...]],
        cumu: int = 0,
        mode: int = 0
        ) -> List[Callable]:
    """
    2025.09.18修改，来源于funs_fd.
    生成一组按数值区间（bin）进行计数的函数，支持滑动区间与累计区间、前闭后开与前开后闭模式。

    📊 典型应用场景：
      - 成绩分组统计（推荐 cumu=0, mode=0）：
          Count[0-60)   → 不及格人数
          Count[60-80)  → 及格人数
          Count[80-100) → 良好人数
          Count[100-inf)→ 优秀人数

      - 名次累计统计（推荐 cumu=1, mode=1）：
          Count(0-60]   → 前60名人数
          Count(0-80]   → 前80名人数
          Count(0-100]  → 前100名人数
          Count(100-inf]→ 100名之后人数

    ⚙️ 参数说明：
      :param thresh: list or tuple
          区间划分阈值列表，例如 [0, 60, 80, 100]。
          - 至少需要2个元素
          - 自动排序（不影响原始数据）
          - 最后一个区间默认延伸至无穷大（inf）

      :param cumu: int, default=0
          控制区间生成方式：
          0 → 滑动区间（互斥分段）: [a0,a1), [a1,a2), [a2,a3), ...
          1 → 固定起点累计 + 最后独立区间: [a0,a1), [a0,a2), ..., [a_last, inf)

      :param mode: int, default=0
          控制区间开闭规则：
          0 → 前闭后开 [a, b) —— 推荐用于成绩、年龄、金额等连续数值
          1 → 前开后闭 (a, b] —— 推荐用于名次、排名、序号等离散序数

    🎯 返回值：
      :return: list of functions
          每个函数接受 array/list 输入，返回满足对应区间的元素个数。
          函数名格式：Count[lower-upper) 或 Count(lower-upper]

    💡 使用示例：
        data = [10, 20, 30, 40, 50, 60, 70, 80, 90, 100, 110]

        # 成绩分组
        counters = make_bin_counters([0,60,80,100], cumu=0, mode=0)
        for c in counters:
            print(f"{c.__name__}: {c(data)}")

        # 名次累计
        counters = make_bin_counters([0,60,80,100], cumu=1, mode=1)
        for c in counters:
            print(f"{c.__name__}: {c(data)}")
    """
    # ========== 参数验证 ==========
    # 1. 验证 thresh 类型
    if not isinstance(thresh, (list, tuple)):
        raise TypeError(f"参数 'thresh' 必须是 list 或 tuple，当前类型: {type(thresh).__name__}")

    # 2. 验证 thresh 长度
    if len(thresh) < 2:
        raise ValueError(f"参数 'thresh' 至少需要2个阈值，当前: {len(thresh)}")

    # 3. 验证 thresh 元素类型和值
    for i, x in enumerate(thresh):
        if not isinstance(x, (int, float)):
            raise TypeError(f"thresh 中的元素必须为数字（int/float），但索引 {i} 的元素是 {type(x).__name__}")
        if math.isnan(x):
            raise ValueError(f"thresh 中的元素不能为 NaN，但索引 {i} 的元素是 NaN")
        if math.isinf(x):
            raise ValueError(f"thresh 中的元素不能为无穷大，但索引 {i} 的元素是 {x}")

    # 4. 验证 cumu 类型和值
    if not isinstance(cumu, int):
        raise TypeError(f"参数 'cumu' 必须是整数，当前类型: {type(cumu).__name__}")
    if cumu not in (0, 1):
        raise ValueError(f"参数 'cumu' 必须是 0 或 1，当前值: {cumu}")

    # 5. 验证 mode 类型和值
    if not isinstance(mode, int):
        raise TypeError(f"参数 'mode' 必须是整数，当前类型: {type(mode).__name__}")
    if mode not in (0, 1):
        raise ValueError(f"参数 'mode' 必须是 0 或 1，当前值: {mode}")

    # 6. 验证 thresh 是否有重复值
    if len(set(thresh)) != len(thresh):
        raise ValueError("thresh 中包含重复值")

    # 排序（不影响原始数据）
    thresh_sorted = sorted(thresh)
    if thresh_sorted != list(thresh):
        print(f"⚠️  警告: thresh 已自动排序（原: {thresh} → 现: {thresh_sorted}）")
    thresh = thresh_sorted

    base = thresh[0]
    intervals = []

    if cumu == 1:
        # 固定起点累计：对每个 thresh[1:] 生成累计区间 [base, t)
        for t in thresh[1:]:
            intervals.append((base, t))
        # 额外追加最后一个独立区间 [last, inf)
        intervals.append((thresh[-1], float('inf')))
    else:  # cumu == 0，默认滑动区间
        thresh_ext = thresh + [float('inf')]
        intervals = [(thresh_ext[i], thresh_ext[i + 1]) for i in range(len(thresh_ext) - 1)]

    def make_counter(lower, upper):
        def counter(arr):
            # 处理 pandas Series 对象
            if isinstance(arr, pd.Series):
                # 检查 Series 是否为空
                if arr.empty:
                    return 0
                # 转换为 numpy 数组
                arr = arr.values
            # 验证输入数据
            elif not isinstance(arr, (list, tuple, np.ndarray)):
                raise TypeError("输入数据必须是列表、元组、numpy 数组或 pandas Series")

            # 转换为 numpy 数组
            arr = np.asarray(arr)

            # 验证数组元素类型
            if not np.issubdtype(arr.dtype, np.number):
                raise TypeError("输入数组中的所有元素必须是数字类型")

            # 过滤掉 NaN 值，只处理有效数值
            valid_arr = arr[~np.isnan(arr)]

            # 检查是否有无穷大值
            if np.any(np.isinf(valid_arr)):
                raise ValueError("输入数组中包含无穷大值")

            # 如果过滤后没有有效数据，返回0
            if len(valid_arr) == 0:
                return 0

            # 根据模式计算计数
            if cumu == 1 and upper != float('inf'):
                # 累计区间部分
                if mode == 0:
                    return np.sum((valid_arr >= lower) & (valid_arr < upper))
                else:
                    return np.sum((valid_arr > lower) & (valid_arr <= upper))
            else:
                # 滑动区间 或 最后独立区间
                if mode == 0:
                    return np.sum((valid_arr >= lower) & (valid_arr < upper))
                else:
                    return np.sum((valid_arr > lower) & (valid_arr <= upper))

        # ========== 函数命名：统一使用 "Count[...)" 或 "Count(...]" ==========
        left_bracket = '[' if mode == 0 else '('
        right_bracket = ')' if mode == 0 else ']'

        # 处理无穷大的显示
        if math.isinf(upper):
            upper_display = "inf"
        else:
            upper_display = upper

        counter.__name__ = f"Count{left_bracket}{lower}-{upper_display}{right_bracket}"
        return counter

    return [make_counter(l, u) for l, u in intervals]

# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 生成一组"双条件计数器"(Dual Condition Counters)，支持名次模式与分数模式。
def make_dual_cond_counters(
        thresh: Union[List[Union[int, float]], Tuple[Union[int, float], ...]],
        sec_thresh: Union[int, float],
        mode: str = 'rank'
        ) -> List[Callable]:
    """
    生成一组"双条件计数器"(Dual Condition Counters)，支持名次模式与分数模式。

    🎯 核心功能：
      - 每个函数统计：主维度落在某区间内 且 次维度满足阈值条件 的记录数
      - 额外统计：次维度不满足阈值条件的记录数（主维度不限）

    📊 典型应用场景：

      🎓 名次模式 (mode='rank')：
        - 统计"学科名次在(0,60] 且 总分名次≤100"的学生
        - 适用于：名次越小越好（如排名）

      📈 分数模式 (mode='score')：
        - 统计"数学分数在[80,90) 且 总分≥450"的学生
        - 适用于：分数越大越好（如成绩、评分、销售额）

    ⚙️ 参数说明：
      :param thresh: list or tuple of numbers
          主维度分段阈值，如 [0, 60, 80, 100] 或 [60, 70, 80, 90]。
          - 自动升序排序（不影响原始数据）
          - 至少需要2个元素
          - 支持 int/float，自动过滤 NaN

      :param sec_thresh: number
          次维度阈值：
          - mode='rank' → 总分名次上限（包含），如 100 → 只统计总分名次 ≤100
          - mode='score' → 总分分数下限（包含），如 450 → 只统计总分分数 ≥450

      :param mode: str, default='rank'
          'rank' → 名次模式：主维度区间 (lower, upper]，次维度条件 ≤ sec_thresh
          'score' → 分数模式：主维度区间 [lower, upper)，次维度条件 ≥ sec_thresh

    🎯 返回值：
      :return: list of functions
          每个函数接受 [(主维度值, 次维度值), ...] 格式数据，返回满足条件的记录数。
          命名格式：
            - 名次模式：'DC(0-60] T≤100'
            - 分数模式：'DC[80-90) T≥450'
            - 超限统计：'DC(0-inf] T>100' 或 'DC[0-inf) T<450'

    💡 使用示例：

        # ===== 名次模式 =====
        data_rank = [(10, 50), (70, 120), (90, 80), (110, 90)]
        counters = make_dual_cond_counters([0,60,80,100], 100, mode='rank')
        for c in counters:
            print(f"{c.__name__}: {c(data_rank)}")

        # ===== 分数模式 =====
        data_score = [(85, 480), (75, 420), (92, 500), (65, 460)]
        counters = make_dual_cond_counters([60,70,80,90], 450, mode='score')
        for c in counters:
            print(f"{c.__name__}: {c(data_score)}")
    """
    # ========== 参数校验 ==========
    # 1. 验证 thresh 类型
    if not isinstance(thresh, (list, tuple)):
        raise TypeError(f"'thresh' 应为 list 或 tuple，当前类型: {type(thresh).__name__}")

    # 2. 验证 thresh 长度
    if len(thresh) < 2:
        raise ValueError(f"'thresh' 至少需要2个阈值，当前: {len(thresh)}")

    # 3. 验证 thresh 元素类型和值
    for i, x in enumerate(thresh):
        if not isinstance(x, (int, float)):
            raise TypeError(f"thresh 中的元素必须为数字（int/float），但索引 {i} 的元素是 {type(x).__name__}")
        if math.isnan(x):
            raise ValueError(f"thresh 中的元素不能为 NaN，但索引 {i} 的元素是 NaN")
        if x < 0:
            raise ValueError(f"thresh 中的元素必须为非负数，但索引 {i} 的元素是 {x}")

    # 4. 验证 sec_thresh 类型和值
    if not isinstance(sec_thresh, (int, float)):
        raise TypeError(f"'sec_thresh' 必须为数字（int/float），当前类型: {type(sec_thresh).__name__}")
    if math.isnan(sec_thresh):
        raise ValueError("'sec_thresh' 不能为 NaN")
    if sec_thresh < 0:
        raise ValueError(f"'sec_thresh' 必须为非负数，当前值: {sec_thresh}")

    # 5. 验证 mode 类型和值
    if not isinstance(mode, str):
        raise TypeError(f"'mode' 必须为字符串，当前类型: {type(mode).__name__}")
    if mode not in ('rank', 'score'):
        raise ValueError(f"mode 必须为 'rank' 或 'score'，当前: {mode}")

    # 6. 验证 thresh 是否为严格升序序列
    thresh_sorted = sorted(thresh)
    if thresh_sorted != list(thresh):
        print(f"⚠️  警告: thresh 已自动排序（原: {thresh} → 现: {thresh_sorted}）")
    thresh = thresh_sorted

    # 检查是否有重复值
    if len(set(thresh)) != len(thresh):
        raise ValueError("thresh 中包含重复值")

    # ========== 根据 mode 设置比较逻辑 ==========
    if mode == 'rank':
        # 名次模式：主维度 (lower, upper]，次维度 <= sec_thresh
        main_lower_op = lambda x, l: x > l  # 严格大于下界
        main_upper_op = lambda x, u: x <= u  # 小于等于上界
        secondary_op = lambda y: y <= sec_thresh
        secondary_fail_op = lambda y: y > sec_thresh
        main_bracket = ('(', ']')
        secondary_prefix = 'T≤'
        secondary_fail_prefix = 'T>'
    else:  # mode == 'score'
        # 分数模式：主维度 [lower, upper)，次维度 >= sec_thresh
        main_lower_op = lambda x, l: x >= l  # 大于等于下界
        main_upper_op = lambda x, u: x < u  # 严格小于上界
        secondary_op = lambda y: y >= sec_thresh
        secondary_fail_op = lambda y: y < sec_thresh
        main_bracket = ('[', ')')
        secondary_prefix = 'T≥'
        secondary_fail_prefix = 'T<'

    # ========== 生成计数器 ==========
    def make_counter(lower, upper):
        def counter(arr):
            # 新增：处理 pandas Series
            if isinstance(arr, pd.Series):
                # 转换为列表
                arr = arr.tolist()

            # 验证输入数据
            if not isinstance(arr, (list, tuple, np.ndarray)):
                raise TypeError("输入数据必须是列表、元组或 numpy 数组")

            count = 0
            for item in arr:
                # 验证每个元素是否为元组且长度为2
                if not isinstance(item, (list, tuple)) or len(item) != 2:
                    raise ValueError("输入数据的每个元素必须是长度为2的元组或列表")

                # 验证元素值是否为数字
                x, y = item
                if not (isinstance(x, (int, float)) and isinstance(y, (int, float))):
                    raise ValueError("输入数据的每个元素必须是数字")

                # 检查条件
                if main_lower_op(x, lower) and main_upper_op(x, upper) and secondary_op(y):
                    count += 1
            return count

        # 格式化函数名
        if math.isinf(upper):
            counter.__name__ = f"DC{main_bracket[0]}{lower}-inf{main_bracket[1]} {secondary_prefix}{sec_thresh}"
        else:
            counter.__name__ = f"DC{main_bracket[0]}{lower}-{upper}{main_bracket[1]} {secondary_prefix}{sec_thresh}"
        return counter

    def last_counter(arr):
        # 新增：处理 pandas Series
        if isinstance(arr, pd.Series):
            # 转换为列表
            arr = arr.tolist()

        # 验证输入数据
        if not isinstance(arr, (list, tuple, np.ndarray)):
            raise TypeError("输入数据必须是列表、元组或 numpy 数组")

        count = 0
        for item in arr:
            # 验证每个元素是否为元组且长度为2
            if not isinstance(item, (list, tuple)) or len(item) != 2:
                raise ValueError("输入数据的每个元素必须是长度为2的元组或列表")

            # 验证元素值是否为数字
            _, y = item
            if not isinstance(y, (int, float)):
                raise ValueError("输入数据的每个元素必须是数字")

            # 检查条件
            if secondary_fail_op(y):
                count += 1
        return count

    last_counter.__name__ = f"DC{main_bracket[0]}0-inf{main_bracket[1]} {secondary_fail_prefix}{sec_thresh}"

    # 添加无穷大，生成区间
    thresh_ext = thresh + [float('inf')]
    intervals = [(thresh_ext[i], thresh_ext[i + 1]) for i in range(len(thresh_ext) - 1)]

    # 生成所有计数器
    counters = [make_counter(l, u) for l, u in intervals] + [last_counter]
    return counters

# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 根据阈值生成一组成绩区间统计函数（计数 + 比率），可选生成低于最低分统计、平均分函数和有效数据个数统计。
def make_rate_counters(
        thresh: Union[List[Union[int, float]], Tuple[Union[int, float], ...], np.ndarray],
        cumu: bool = True,
        include_mean: bool = True,
        include_below_min: bool = False,
        include_count_valid: int = 0
) -> Tuple[Callable, ...]:
    """
    根据阈值生成一组成绩区间统计函数（计数 + 比率），可选生成低于最低分统计、平均分函数和有效数据个数统计。
    2025.10.20,阿里灵码优化。
    🎯 核心特性：
      - 直接返回函数组，无需字典包装；
      - 函数名采用数学区间风格：count[60,80)、ratio[60,80)、count[80,+∞)、ratio[80,+∞)；
      - 支持区间统计和累计统计两种模式；
      - 可选生成低于最低分的统计函数（在两种模式下均有效）；
      - 可选生成平均分计算函数（自动忽略NaN值）；
      - 新增：可选生成有效数据个数统计函数（count_valid），位置由 include_count_valid 控制。

    📊 位置控制逻辑：
      - include_count_valid = 0: 不添加 count_valid；
      - include_count_valid = 1: 将 count_valid 添加在函数列表的第一个位置；
      - include_count_valid = -1: 将 count_valid 添加在函数列表的最后一个位置。

    ⚠️ 注意事项：
      - 当 cumu=True 时，区间表示分数 >= 阈值的比例，例如 ratio[80,+∞) 表示 80 分及以上占比。
      - 比率计算时，分母为输入数组的总长度（包括 NaN 值）。若需基于有效数据计算，请先过滤 NaN。

    :param thresh: list/tuple/ndarray，阈值列表，必须是长度 >=2 的升序序列；
    :param cumu: bool，是否为累计统计模式。True 时，每个阈值 t 生成 [t, +∞) 区间统计（即 >=t）。
    :param include_mean: bool, 是否生成计算平均分的函数，默认 True；
    :param include_below_min: bool, 是否生成低于最低分的统计函数，默认 False；
    :param include_count_valid: int, 控制 count_valid 的位置（0=不添加, 1=第一个位置, -1=最后一个位置）；
    :return: tuple，包含所有生成的统计函数。

    :raises ValueError: 当参数不符合要求时；
    :raises TypeError: 当参数类型不正确时。
    """
    # 参数验证。
    if not isinstance(thresh, (list, tuple, np.ndarray)):
        raise TypeError("thresh 必须为列表、元组或 numpy 数组。")

    if isinstance(thresh, np.ndarray):
        thresh = thresh.tolist()

    if len(thresh) < 2:
        raise ValueError("thresh 必须包含至少 2 个元素。")

    for i, t in enumerate(thresh):
        if not isinstance(t, (int, float)):
            raise TypeError(f"thresh 中的所有元素必须是数字类型，但索引 {i} 的元素是 {type(t)}。")

    for i in range(len(thresh) - 1):
        if thresh[i] >= thresh[i + 1]:
            raise ValueError("thresh 必须是严格升序序列。")

    if not isinstance(cumu, bool):
        raise TypeError("cumu 必须是布尔值。")

    if not isinstance(include_below_min, bool):
        raise TypeError("include_below_min 必须是布尔值。")

    if not isinstance(include_mean, bool):
        raise TypeError("include_mean 必须是布尔值。")

    if not isinstance(include_count_valid, int):
        raise TypeError("include_count_valid 必须是整数（0, 1, -1）。")

    if include_count_valid not in [0, 1, -1]:
        raise ValueError("include_count_valid 必须是 0, 1 或 -1。")

    for i, t in enumerate(thresh):
        if t < 0:
            raise ValueError(f"thresh 中的所有元素必须为非负数，但索引 {i} 的元素为 {t}。")

    def make_threshold_func(lower, upper=None, ratio=False, cumu_mode=False, is_last_interval=False, below_min=False):
        """
        内部辅助函数，用于创建阈值相关的统计函数（计数/比率）。
        """

        def func(scores):
            # 处理 pandas Series 对象。
            if isinstance(scores, pd.Series):
                if scores.empty:
                    return np.nan if ratio else 0
                arr = scores.values
            elif not isinstance(scores, (list, tuple, np.ndarray)):
                raise TypeError("输入分数必须是列表、元组或 numpy 数组。")
            else:
                arr = np.array(scores)

            # 尝试将数组转换为 float，以处理混合类型或 NaN
            try:
                arr = arr.astype(float)
            except (ValueError, TypeError):
                raise TypeError("分数数组必须能转换为数值类型。")

            mask = np.zeros(arr.shape, dtype=bool)  # 初始化掩码。

            if below_min:
                mask = arr < lower
            elif cumu_mode:
                # 在 cumu 模式下，is_last_interval 参数不再传入，逻辑简化
                mask = arr >= lower
            else:
                if is_last_interval:
                    mask = (arr >= lower) & (arr <= upper)
                else:
                    mask = (arr >= lower) & (arr < upper)

            count = np.sum(mask)
            total = len(arr)
            return count / total if ratio and total > 0 else int(count)

        # 生成函数名。
        prefix = "ratio" if ratio else "count"
        if below_min:
            func_name = f"{prefix}(-∞,{lower})"
        elif cumu_mode:
            func_name = f"{prefix}[{lower},+∞)"
        else:
            if is_last_interval:
                func_name = f"{prefix}[{lower},{upper}]"
            else:
                func_name = f"{prefix}[{lower},{upper})"
        func.__name__ = func_name
        return func

    def make_mean_func():
        """
        平均分计算函数：自动忽略 NaN 值。
        """

        def func(scores):
            if isinstance(scores, pd.Series):
                arr = scores.values
            elif not isinstance(scores, (list, tuple, np.ndarray)):
                raise TypeError("输入分数必须是列表、元组或 numpy 数组。")
            else:
                arr = np.array(scores)

            # 尝试将数组转换为 float
            try:
                arr = arr.astype(float)
            except (ValueError, TypeError):
                raise TypeError("分数数组必须能转换为数值类型。")

            # 过滤掉 NaN 值。
            valid = arr[~np.isnan(arr)]

            if len(valid) == 0:
                return np.nan

            return float(np.mean(valid))

        func.__name__ = "mean"
        return func

    def make_count_valid_func():
        """
        有效数据个数统计函数。
        """

        def func(scores):
            if isinstance(scores, pd.Series):
                arr = scores.values
            elif not isinstance(scores, (list, tuple, np.ndarray)):
                raise TypeError("输入分数必须是列表、元组或 numpy 数组。")
            else:
                arr = np.array(scores)

            # 尝试将数组转换为 float
            try:
                arr = arr.astype(float)
            except (ValueError, TypeError):
                raise TypeError("分数数组必须能转换为数值类型。")

            # 计算有效数据个数（非 NaN）。
            valid = ~np.isnan(arr)
            return int(np.sum(valid))

        func.__name__ = "count_valid"
        return func

    funcs = []
    n = len(thresh)
    min_thresh = thresh[0]

    # 生成低于最低分的统计。
    if include_below_min:
        below_count_func = make_threshold_func(min_thresh, ratio=False, below_min=True)
        below_ratio_func = make_threshold_func(min_thresh, ratio=True, below_min=True)
        funcs.extend([below_count_func, below_ratio_func])

    # 生成阈值区间统计。
    for i in range(n - 1):
        lower = thresh[i]
        upper = thresh[i + 1]
        is_last_interval = (i == n - 2)

        # 根据 cumu 模式选择调用方式，避免传递无用参数
        if cumu:
            count_func = make_threshold_func(lower, ratio=False, cumu_mode=True)
            ratio_func = make_threshold_func(lower, ratio=True, cumu_mode=True)
        else:
            count_func = make_threshold_func(lower, upper, ratio=False, is_last_interval=is_last_interval)
            ratio_func = make_threshold_func(lower, upper, ratio=True, is_last_interval=is_last_interval)

        funcs.extend([count_func, ratio_func])

    # 生成平均分统计。
    if include_mean:
        mean_func = make_mean_func()
        funcs.append(mean_func)

    # 根据 include_count_valid 添加有效数据个数统计。
    if include_count_valid != 0:
        count_valid_func = make_count_valid_func()
        if include_count_valid == 1:
            funcs.insert(0, count_valid_func)  # 插入到第一个位置。
        elif include_count_valid == -1:
            funcs.append(count_valid_func)  # 插入到最后一个位置。

    return tuple(funcs)


# --- 示例用法 ---
if __name__ == "__main__":
    # 定义阈值
    thresholds = [60, 70, 80, 90]
    print(f"阈值: {thresholds}")
    print(f"累计模式 (cumu=True):")

    # 生成统计函数组
    stat_funcs = make_rate_counters(thresholds, cumu=True, include_mean=True, include_below_min=True,
                                    include_count_valid=-1)

    # 示例数据
    sample_scores = [55, 65, 75, 85, 95, 88, 72, 60, 90, np.nan]
    print(f"示例数据: {sample_scores}")

    # 打印所有生成的函数名称
    print("\n生成的函数列表:")
    for i, func in enumerate(stat_funcs):
        print(f"  [{i:2d}] {func.__name__}")

    print("\n计算示例数据的统计结果:")
    for func in stat_funcs:
        result = func(sample_scores)
        print(f"  {func.__name__}: {result:.4f}" if isinstance(result, float) else f"  {func.__name__}: {result}")

    print("\n" + "=" * 50)
    print(f"非累计模式 (cumu=False):")
    stat_funcs2 = make_rate_counters(thresholds, cumu=False, include_mean=True, include_below_min=True,
                                     include_count_valid=-1)

    print("\n生成的函数列表:")
    for i, func in enumerate(stat_funcs2):
        print(f"  [{i:2d}] {func.__name__}")

    print("\n计算示例数据的统计结果:")
    for func in stat_funcs2:
        result = func(sample_scores)
        print(f"  {func.__name__}: {result:.4f}" if isinstance(result, float) else f"  {func.__name__}: {result}")


# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■


# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■   df数据与工作表单元  ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 返回工作薄与指定工作表。
def trim_wb(
        wb: Workbook,
        sheet_names: Union[str, List[str]]
        ) -> Tuple[Workbook, Union[Worksheet, List[Worksheet]]]:
    """
    只保留指定名称的工作表，并删除工作簿中的其他所有工作表。

    :param wb: openpyxl工作簿对象
    :param sheet_names: 要保留的工作表名称（可以是字符串或字符串列表）
    :return: 元组包含（修改后的工作簿对象, 保留的工作表对象/列表）
    :raises ValueError: 如果指定的工作表名称不存在于工作簿中
    """
    # 统一输入格式为列表
    if isinstance(sheet_names, str):
        sheet_names = [sheet_names]

    # 检查所有指定的工作表是否存在
    missing_sheets = [name for name in sheet_names if name not in wb.sheetnames]
    if missing_sheets:
        raise ValueError(f"工作表 {missing_sheets} 不存在于工作簿中")

    # 删除不在保留列表中的工作表
    sheets_to_remove = [sheet for sheet in wb.worksheets if sheet.title not in sheet_names]
    for sheet in sheets_to_remove:
        wb.remove(sheet)

    # 返回工作簿和保留的工作表对象
    retained_sheets = [wb[name] for name in sheet_names]

    # 如果只保留一个工作表，返回单个工作表对象；否则返回列表
    if len(sheet_names) == 1:
        return wb, retained_sheets[0]
    else:
        return wb, retained_sheets

# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 将一个或多个Pandas DataFrame逐一注入到一个由openpyxl创建的工作表中。
def dfs_to_ws(
        ws: Any,
        row: int,
        col: int,
        dfs: Union[pd.DataFrame, List[pd.DataFrame]],
        rg: int = 10,
        cg: int = 0,
        hd: bool = False,
        idx: bool = False,
        na_rep: Optional[Any] = None
        ) -> None:
    """
    将一个或多个Pandas DataFrame逐一注入到一个由openpyxl创建的工作表中。
    专注于核心功能：数据写入和布局控制。

    参数:
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        由openpyxl创建的工作表对象，数据将被写入此工作表

    row : int
        第一个DataFrame的起始行坐标（从1开始计数）

    col : int
        第一个DataFrame的起始列坐标（从1开始计数）

    dfs : pandas.DataFrame 或 list of pandas.DataFrame
        要写入的一个或多个DataFrame对象。可以是单个DataFrame或DataFrame列表

    rg : int, 可选, 默认值: 10
        DataFrame之间的行间距（行间隔数）。控制一个DataFrame结束后到下一个
        DataFrame开始前的空行数量

    cg : int, 可选, 默认值: 0
        DataFrame之间的列间距（列间隔数）。控制一个DataFrame结束后到下一个
        DataFrame开始前的空列数量

    hd : bool, 可选, 默认值: False
        是否包含DataFrame的列头（表头）。如果为True，则将DataFrame的列名
        作为第一行写入

    idx : bool, 可选, 默认值: False
        是否包含DataFrame的索引。如果为True，则将索引作为第一列写入

    na_rep : 任意类型, 可选, 默认值: None
        NaN值的替代表示。当DataFrame中包含NaN、NaT等空值时，使用此值进行替换。
        如果保持为None，则空值将保持为None（在Excel中显示为空单元格）

    返回:
    -------
    None
        此函数不返回任何值，直接修改传入的工作表对象

    异常:
    ------
    ValueError
        如果ws参数不是有效的Worksheet对象，或者dfs参数不是DataFrame或DataFrame列表

    注意事项:
    ---------
    1. 此函数会直接修改传入的工作表对象，但不会自动保存工作簿
    2. 空值处理使用Pandas的isnull()方法，可以识别多种空值类型（NaN、NaT等）
    3. 对于大型DataFrame，建议使用na_rep参数处理空值，避免Excel显示错误
    4. 函数采用批量写入方式优化性能，减少方法调用次数
    """
    # 检查ws是否为Worksheet对象
    if not hasattr(ws, 'cell'):
        raise ValueError("ws必须是一个openpyxl的Worksheet对象")

    # 将单个DataFrame转换为列表
    if isinstance(dfs, pd.DataFrame):
        dfs = [dfs]

    # 检查dfs是否为DataFrame列表
    if not all(isinstance(df, pd.DataFrame) for df in dfs):
        raise ValueError("dfs必须是一个Pandas DataFrame对象或包含Pandas DataFrame对象的列表")

    # 检查行列参数的有效性
    if row < 1 or col < 1:
        raise ValueError("行和列参数必须大于等于1")

    # 遍历每个DataFrame
    for df_idx, df in enumerate(dfs):
        # 检查DataFrame是否为空
        if df.empty:
            warnings.warn(f"第{df_idx + 1}个DataFrame为空，跳过处理")
            continue

        # 获取所有行数据（提前转换为列表，避免重复生成）
        try:
            rows = list(dataframe_to_rows(df, index=idx, header=hd))
        except Exception as e:
            raise ValueError(f"处理第{df_idx + 1}个DataFrame时出错: {str(e)}")

        # 批量写入数据
        for r_offset, row_data in enumerate(rows):
            for c_offset, value in enumerate(row_data):
                # 处理NaN值
                if pd.isnull(value):
                    value = na_rep
                # 处理元组类型的值（如多级列索引名称）
                elif isinstance(value, tuple):
                    # 将元组转换为字符串形式
                    value = '_'.join(str(v) for v in value)

                # 计算实际单元格位置
                current_row = row + r_offset
                current_col = col + c_offset

                # 检查单元格是否为合并单元格，如果是，则找到合并区域的起始单元格
                cell = ws.cell(row=current_row, column=current_col)
                if hasattr(cell, 'merged_cell') and cell.merged_cell:
                    # 对于合并单元格，我们需要找到左上角的单元格来写入值
                    for merged_range in ws.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            # 获取合并区域的起始单元格
                            cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                            break
                
                # 直接赋值
                cell.value = value

        # 更新位置为下一个DataFrame的起始位置
        row +=  rg
        col +=  cg

# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 检查输入数据是否为 pandas DataFrame，并验证以下条件
def validate_dataframe(
    data,
    required_columns: list,
    allow_extra_columns: bool = True
    ) -> pd.DataFrame:
    """
    检查输入数据是否为 pandas DataFrame，并验证数据结构与内容质量。

    检查项包括：
    1. 是否为 DataFrame；
    2. 是否包含所有必需列；
    3. （可选）是否包含额外列；
    4. 哪些列存在空值（NaN/None）；
    5. 哪些数值列存在 0 值。

    返回一个包含中文检查项、通过状态和详细说明的反馈表格。
    """
    results = []

    # 1. 是否为 DataFrame
    is_dataframe = isinstance(data, pd.DataFrame)
    results.append({
        "检查项": "数据类型",
        "是否通过": is_dataframe,
        "说明": "输入是 pandas DataFrame" if is_dataframe else "输入不是 pandas DataFrame"
    })

    if not is_dataframe:
        return pd.DataFrame(results)

    df = data
    if df.empty:
        results.append({
            "检查项": "数据是否为空",
            "是否通过": False,
            "说明": "DataFrame 为空（无行或无列）"
        })
        return pd.DataFrame(results)

    actual_columns = set(df.columns)
    required_set = set(required_columns)

    # 2. 是否包含所有必需列
    missing_required = sorted(required_set - actual_columns)
    has_all_required = len(missing_required) == 0
    msg_req = "包含所有必需列" if has_all_required else f"缺少必需列: {missing_required}"
    results.append({
        "检查项": "必需列标",
        "是否通过": has_all_required,
        "说明": msg_req
    })

    # 3. 是否有额外列（仅当不允许额外列时检查）
    if not allow_extra_columns:
        extra_cols = sorted(actual_columns - required_set)
        no_extra = len(extra_cols) == 0
        msg_extra = "无额外列（列结构严格匹配）" if no_extra else f"存在额外列: {extra_cols}"
        results.append({
            "检查项": "额外列标",
            "是否通过": no_extra,
            "说明": msg_extra
        })

    # 4. 检查空值（NaN/None）——列出具体列
    nan_series = df.isnull().any()
    nan_cols = sorted(nan_series[nan_series].index.tolist())
    has_nan = len(nan_cols) > 0
    msg_nan = "无空值" if not has_nan else f"以下列包含空值: {nan_cols}"
    results.append({
        "检查项": "检查空值",
        "是否通过": not has_nan,
        "说明": msg_nan
    })

    # 5. 检查 0 值（仅数值列）——列出具体列
    numeric_cols = df.select_dtypes(include=[np.number]).columns
    zero_cols = []
    if len(numeric_cols) > 0:
        zero_mask = (df[numeric_cols] == 0)
        zero_cols = sorted(zero_mask.any()[zero_mask.any()].index.tolist())
    has_zero = len(zero_cols) > 0
    msg_zero = "无数值列中的 0 值" if not has_zero else f"以下列包含0值: {zero_cols}"
    results.append({
        "检查项": "检查0值",
        "是否通过": not has_zero,
        "说明": msg_zero
    })

    return pd.DataFrame(results)

# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# ================================df数据操作单元=====================================================
# 将具有多级列索引的DataFrame按照指定层级拆分成多个DataFrame。
def df_split_levels(
        df: pd.DataFrame,
        level: int = 0,
        keep_level: Union[int, bool] = 0
        ) -> Dict[Any, pd.DataFrame]:
    """
    将具有多级列索引的DataFrame按照指定层级拆分成多个DataFrame。

    参数:
        df: 输入的DataFrame，应该具有多级列索引
        level: 拆分DataFrame的层级，默认为0（第一级）
        keep_level:
            - 0: 删除拆分所用的层级（默认）
            - 1: 保留所有层级
            - -1: 将当前层级的索引值添加到下一级索引值中，然后删除当前层级

    返回:
        字典{层级值: DataFrame}，其中键是拆分层级的唯一值

    异常:
        ValueError: 如果输入的DataFrame没有多级列索引或指定的层级无效
    """
    # 确保DataFrame具有多级列索引
    if not isinstance(df.columns, pd.MultiIndex):
        raise ValueError("输入的DataFrame必须具有多级列索引")

    # 验证指定的层级是否有效
    nlevels = df.columns.nlevels
    if not (-nlevels <= level < nlevels):
        raise ValueError(f"层级 {level} 超出了DataFrame的列索引范围 (-{nlevels} 到 {nlevels - 1})")

    # 验证keey_level参数
    if keep_level not in [0, 1, -1]:
        raise ValueError("keey_level参数必须是0、1或-1")

    # 处理负索引
    level = level % nlevels

    # 获取指定层级的所有唯一标签
    level_values = df.columns.get_level_values(level)
    unique_labels = level_values.unique()

    # 存储结果
    result = {}

    # 遍历每个唯一标签
    for label in unique_labels:
        # 选择所有在指定层级具有该标签的列
        mask = level_values == label
        temp_df = df.loc[:, mask].copy()

        # 处理列索引
        if keep_level == 0:
            # 删除拆分所用的层级
            if temp_df.columns.nlevels > 1:
                temp_df.columns = temp_df.columns.droplevel(level)

        elif keep_level == -1:
            # 将当前层级的索引值添加到下一级索引值中
            if temp_df.columns.nlevels > 1:
                # 构建新的列名
                new_columns = []
                for col in temp_df.columns:
                    # 将元组转换为列表
                    col_list = list(col)

                    # 构建新的列名
                    if level + 1 < len(col_list):
                        # 合并当前层级和下一层级的值
                        merged_value = f"{col_list[level]}_{col_list[level + 1]}"
                        # 删除当前层级和下一层级
                        del col_list[level:level + 2]
                        # 插入合并后的值
                        col_list.insert(level, merged_value)
                    else:
                        # 只有当前层级的情况
                        merged_value = str(col_list[level])
                        # 删除当前层级
                        del col_list[level]
                        col_list.append(merged_value)

                    # 如果只剩下一个层级，则直接使用该值
                    new_columns.append(col_list[0] if len(col_list) == 1 else tuple(col_list))

                temp_df.columns = new_columns

        # 无论keey_level为何值，都不显示列索引名称
        if isinstance(temp_df.columns, pd.MultiIndex):
            # 对于多级索引，将所有层级的名称设置为None
            temp_df.columns.names = [None] * temp_df.columns.nlevels
        else:
            # 对于单级索引，将名称设置为None
            temp_df.columns.name = None

        result[label] = temp_df

    return result

# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 按指定列将 DataFrame 分割成多个子 DataFrame，并将它们存储在一个字典中。
def df_groupby_col(
        df: pd.DataFrame,
        col: str,
        sort_groups: Union[bool, List[Any], Callable, str] = False,
        sort_ascending: bool = True) -> Dict[Any, pd.DataFrame]:
    """
    按指定列将 DataFrame 分割成多个子 DataFrame，并将它们存储在一个字典中。
    支持多种方式对分组进行排序。

    :param df: DataFrame 对象
    :param col: 用于分组的列标签
    :param sort_groups: 控制分组排序的方式
        - False: 不排序，保持原始出现顺序 (默认)
        - True: 按组名排序
        - List: 按提供的列表顺序排序
        - Callable: 使用自定义函数对组名排序
        - str: 按指定列的汇总值排序 (如 'sum', 'mean', 'count', 'min', 'max', 'std')
    :param sort_ascending: 排序顺序，True 为升序，False 为降序
    :return: 一个字典，键为组名，值为对应的子 DataFrame（索引已重置）
    """
    # 输入校验
    if not isinstance(df, pd.DataFrame):
        raise TypeError(f"参数 'df' 必须是 pandas.DataFrame，当前类型: {type(df).__name__}")
    if col not in df.columns:
        raise KeyError(f"列 '{col}' 不存在于 DataFrame 中。可用列: {list(df.columns)}")
    if df.empty:
        return {}

    # 使用 groupby 获取分组，sort=False 保持原始首次出现顺序
    grouped = df.groupby(col, sort=False)
    groups = {name: group.reset_index(drop=True) for name, group in grouped}

    # 若无需排序，直接返回
    if not sort_groups:
        return groups

    # 获取排序后的键列表
    if isinstance(sort_groups, list):
        # 按用户指定列表排序，缺失的键追加在末尾
        sorted_keys = [k for k in sort_groups if k in groups] + [k for k in groups if k not in sort_groups]
    elif callable(sort_groups):
        # 使用自定义函数排序
        sorted_keys = sorted(groups.keys(), key=sort_groups, reverse=not sort_ascending)
    elif isinstance(sort_groups, str) and sort_groups in ['sum', 'mean', 'count', 'min', 'max', 'std']:
        # 按数值列聚合值排序（优先使用数值列，避免对字符串列求sum/mean）
        numeric_cols = df.select_dtypes(include=['number']).columns.drop(col, errors='ignore')
        if len(numeric_cols) == 0:
            # 无数字列 → 退化为按键名排序
            sorted_keys = sorted(groups.keys(), reverse=not sort_ascending)
        else:
            # 对每个分组的数值列聚合后求和，作为排序依据
            agg_series = df.groupby(col)[numeric_cols].agg(sort_groups).sum(axis=1)
            sorted_keys = agg_series.sort_values(ascending=sort_ascending).index.tolist()
    else:
        # sort_groups 为 True 或其他 Truthy 值 → 按键名排序
        sorted_keys = sorted(groups.keys(), reverse=not sort_ascending)

    # 按排序后的键重建字典（Python 3.7+ 保持插入顺序）
    return {key: groups[key] for key in sorted_keys}

# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 按列索引列表将 DataFrame 分割成多个子 DataFrame，并返回一个字典
def df_split_col(
        df: pd.DataFrame,
        col: List[str],
        include_other: bool = True) -> Dict[str, pd.DataFrame]:
    """
    按列索引列表将 DataFrame 分割成多个子 DataFrame，并返回一个字典

    参数:
    -----------
    df : pd.DataFrame
        原始数据框，待拆分对象
    col : List[str]
        需要拆分的列名列表
    include_other : bool, 可选
        是否在每个子数据框中保留“非拆分列”的其他列，默认为 True

    返回:
    --------
    Dict[str, pd.DataFrame]
        字典，键为拆分列名，值为对应的子数据框
    """
    # 过滤出实际存在于 DataFrame 中的列
    existing_cols = [c for c in col if c in df.columns]

    # 如果没有有效列，返回空字典
    if not existing_cols:
        return {}

    # 初始化结果字典
    result = {}

    if include_other:
        # 获取“其他列”（即不在拆分列表中的列）
        other_cols = [c for c in df.columns if c not in col]

        # 预先复制一份“其他列”的数据框，避免在循环中重复切片
        other_df = df[other_cols].copy()

        # 为每个存在的拆分列创建子数据框
        for c in existing_cols:
            # 使用 assign 动态添加当前列，避免多次复制“其他列”数据
            result[c] = other_df.assign(**{c: df[c]})
    else:
        # 仅包含指定的单列
        for c in existing_cols:
            result[c] = df[[c]].copy()

    return result

# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
#　将指定列的数据合并为新列，并返回新DF
def df_comb_cols(
        df: pd.DataFrame,
        cols: List[str],
        name: str,
        combine_func: Union[str, Callable] = 'tuple',
        drop_original: bool = False,
        handle_na: str = 'keep') -> pd.DataFrame:
    '''
    将指定列的数据合并为新列，并返回新DF

    Parameters:
    -----------
    df : pd.DataFrame
        输入的数据框，包含需要处理的数据
    cols : List[str]
        需要处理的列名列表
    name : str
        处理后生成的新列的名称
    combine_func : Union[str, Callable], optional
        合并函数，可以是以下选项：
        - 'tuple': 将值组合为元组 (默认)
        - 'list': 将值组合为列表
        - 'str': 将值组合为字符串（用空格分隔）
        - Callable: 自定义合并函数
    drop_original : bool, optional
        是否删除原始列，默认为False
    handle_na : str, optional
        处理空值的方式，可以是以下选项：
        - 'keep': 保留空值 (默认)
        - 'skip': 跳过包含空值的行
        - 'fill': 用指定值填充空值

    Returns:
    --------
    pd.DataFrame
        处理后的数据框

    Raises:
    -------
    TypeError
        如果参数类型不正确
    ValueError
        如果参数值不正确或列不存在
    '''
    # 参数验证
    if not isinstance(df, pd.DataFrame):
        raise TypeError("df必须是pandas DataFrame类型")

    if not isinstance(cols, list):
        raise TypeError("cols必须是列表类型")

    if not cols:
        raise ValueError("cols不能为空")

    if not isinstance(name, str) or not name:
        raise ValueError("name必须是非空字符串")

    if name in df.columns:
        raise ValueError(f"列名 '{name}' 已存在于DataFrame中")

    # 检查列是否存在
    missing_cols = [col for col in cols if col not in df.columns]
    if missing_cols:
        raise ValueError(f"列 {missing_cols} 在DataFrame中未找到")

    # 创建数据副本，避免修改原始数据
    result_df = df.copy()

    # 处理空值
    if handle_na == 'skip':
        # 跳过包含空值的行
        result_df = result_df.dropna(subset=cols)
    elif handle_na == 'fill':
        # 用空字符串填充空值
        result_df[cols] = result_df[cols].fillna('')

    # 根据combine_func参数选择合并方式
    if combine_func == 'tuple':
        result_df[name] = list(zip(*[result_df[col] for col in cols]))
    elif combine_func == 'list':
        result_df[name] = [list(x) for x in zip(*[result_df[col] for col in cols])]
    elif combine_func == 'str':
        result_df[name] = [' '.join(map(str, x)) for x in zip(*[result_df[col] for col in cols])]
    elif callable(combine_func):
        result_df[name] = [combine_func(*x) for x in zip(*[result_df[col] for col in cols])]
    else:
        raise ValueError(f"不支持的combine_func: {combine_func}")

    # 是否删除原始列
    if drop_original:
        result_df = result_df.drop(columns=cols)

    return result_df

# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 将指定列(cols)中的每一列与目标列(target_col)两两组合，并返回DataFrame。
def df_pair_cols(
        df: pd.DataFrame,
        cols: List[str],
        target_col: str,
        func: Optional[Callable] = None,
        inplace: bool = False,
        drop_target: bool = False) -> pd.DataFrame:
    """
    将指定列(cols)中的每一列与目标列(target_col)两两组合，并返回DataFrame。

    默认情况下，将两列值组合为元组，但可以通过func参数自定义组合方式。

    :param df: 输入的pandas DataFrame
    :param cols: 需要与目标列组合的列名列表
    :param target_col: 目标列名，与cols中的列进行组合
    :param func: 可选的组合函数，应接受两个参数并返回一个值
                如果为None，则使用默认的元组组合方式
    :param inplace: 是否原地修改DataFrame，默认为False
    :param drop_target: 是否在组合后删除目标列，默认为False
    :return: 处理后的DataFrame
    :raises TypeError: 如果参数类型不正确
    :raises ValueError: 如果target_col为空或不存在
    """
    # 参数验证
    if not isinstance(df, pd.DataFrame):
        raise TypeError("df 必须是 pandas DataFrame")

    if not isinstance(cols, list):
        raise TypeError("cols 必须是列表")

    if not isinstance(target_col, str) or not target_col.strip():
        raise ValueError("target_col 必须是非空字符串")

    if func is not None and not callable(func):
        raise TypeError("func 必须是可调用对象或None")

    # 检查目标列是否存在
    if target_col not in df.columns:
        raise ValueError(f"目标列 '{target_col}' 在DataFrame中不存在")

    # 决定是否创建副本
    if not inplace:
        df = df.copy()

    # 过滤出存在的列
    valid_cols = [col for col in cols if col in df.columns]
    missing_cols = [col for col in cols if col not in df.columns]

    # 发出警告，提示缺失的列
    if missing_cols:
        warnings.warn(f"以下列在DataFrame中不存在，将被跳过: {missing_cols}")

    # 如果没有有效的列，直接返回
    if not valid_cols:
        warnings.warn("没有有效的列需要组合")
        return df

    # 组合列
    for col in valid_cols:
        if func is None:
            # 默认行为：将两列组合为元组
            df[col] = list(zip(df[col], df[target_col]))
        else:
            # 使用自定义函数组合列
            df[col] = [func(x, y) for x, y in zip(df[col], df[target_col])]

    # 如果需要，删除目标列
    if drop_target and target_col in df.columns:
        df.drop(columns=[target_col], inplace=True)

    return df


# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 给一个DataFrame的某此列添加排列。
def df_add_cols_rank(df: pd.DataFrame, columns_to_rank: list, ascending=False) -> pd.DataFrame:
    """
    给一个DataFrame的某此列添加排列。

    排名使用 pd.Series.rank() 方法，可以处理 NaN 值和并列情况。
    默认使用降序排名（分数/比率越高，排名越靠前）。

    :param df: 输入的 pandas DataFrame。
    :param columns_to_rank: 需要添加排名的列标识列表。可以是列名（str）或列序号（int）的列表。
                            例如 ["ratio[60,80)", "mean"] 或 [1, 4] 或 ["ratio[60,80)", 4]。
    :param ascending: bool, 是否升序排名。False (默认) 表示数值越大排名越靠前（如高分排名靠前），
                      True 表示数值越小排名越靠前（如低错误率排名靠前）。
    :return: 返回一个新的 DataFrame，其中在指定列后添加了排名列。
    """
    # 防止修改原始 DataFrame
    new_df = df.copy()

    # 首先将所有输入转换为列名
    resolved_col_names = []
    for item in columns_to_rank:
        if isinstance(item, str):
            # 如果是字符串，直接作为列名
            if item not in new_df.columns:
                print(f"警告: 指定的列名 '{item}' 在 DataFrame 中不存在，将跳过。")
                continue
            resolved_col_names.append(item)
        elif isinstance(item, int):
            # 如果是整数，检查是否为有效的列索引
            if item < 0:
                item = len(new_df.columns) + item # 处理负索引
            if 0 <= item < len(new_df.columns):
                col_name = new_df.columns[item]
                resolved_col_names.append(col_name)
            else:
                print(f"警告: 指定的列序号 {item} 超出范围 [0, {len(new_df.columns)-1}]，将跳过。")
                continue
        else:
            print(f"警告: 列标识 '{item}' 类型无效 (应为 str 或 int)，将跳过。")
            continue

    # 从后往前遍历，以避免列索引因插入新列而变化
    for col_name in reversed(resolved_col_names):
        source_series = new_df[col_name]

        # 计算排名
        # method='min' 表示并列项目取最小排名 (例如，两个第一，则下一个为第三)
        # na_option='keep' 表示 NaN 值排名为 NaN
        # ascending=False 表示数值大的排名靠前 (1, 2, 3...)
        ranks = source_series.rank(method='min', na_option='keep', ascending=ascending)

        # 找到源列的索引位置
        source_col_idx = new_df.columns.get_loc(col_name)

        # 将排名列插入到源列之后
        rank_col_name = f"{col_name}_rank"
        new_df.insert(loc=source_col_idx + 1, column=rank_col_name, value=ranks)

    return new_df





# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 合并多个DataFrame，处理重复列名并提供详细的警告和错误信息。
def merge_mult_dfs(
        dfs: List[pd.DataFrame],
        on: Union[str, List[str]],
        how: str = 'outer',
        keep_last: bool = True) -> pd.DataFrame:
    """
    合并多个DataFrame，处理重复列名并提供详细的警告和错误信息。

    参数:
    dfs: DataFrame列表，需要合并的所有DataFrame
    on: 字符串或列表，合并的键列
    how: 合并方式，可选 'left', 'right', 'outer', 'inner'，默认为'outer'
    keep_last: 布尔值，当有重复列时是否保留最后一个DataFrame的值，默认为True

    返回:
    合并后的DataFrame

    异常:
    ValueError: 当输入参数无效或DataFrame缺少必要的键列时
    TypeError: 当输入参数类型不正确时

    警告:
    - 当合并键的数据类型不一致时
    - 当有重复列需要处理时
    - 当合并可能产生意外结果时（如how='left'或how='right'）
    """
    # 验证输入
    if not dfs:
        warnings.warn("提供的DataFrame列表为空，返回空DataFrame")
        return pd.DataFrame()

    if not isinstance(dfs, list):
        raise TypeError("dfs 必须是DataFrame列表")

    if len(dfs) == 1:
        warnings.warn("只提供了一个DataFrame，直接返回其副本")
        return dfs[0].copy()

    # 确保on是列表形式
    if isinstance(on, str):
        on_keys = [on]
    else:
        on_keys = list(on)  # 确保是可变的列表

    # 验证所有DataFrame都包含必要的键列
    for i, df in enumerate(dfs):
        if not isinstance(df, pd.DataFrame):
            raise TypeError(f"dfs[{i}] 不是pandas DataFrame")

        missing_keys = [key for key in on_keys if key not in df.columns]
        if missing_keys:
            raise ValueError(f"DataFrame {i} 缺少键列: {missing_keys}")

    # 检查合并键的数据类型一致性
    for key in on_keys:
        dtypes = []
        for i, df in enumerate(dfs):
            dtype = str(df[key].dtype)
            dtypes.append((i, dtype))

        # 检查所有DataFrame中同一键的数据类型是否一致
        unique_dtypes = set(dtype for _, dtype in dtypes)
        if len(unique_dtypes) > 1:
            dtype_info = ", ".join([f"df{i}: {dtype}" for i, dtype in dtypes])
            warnings.warn(
                f"合并键 '{key}' 的数据类型在不同DataFrame中不一致: {dtype_info}. "
                "这可能导致合并错误或意外结果。"
            )

    # 使用reduce逐步合并
    def merge_two_dfs(df_left, df_right):
        # 找出除了连接键以外的重复列名
        common_cols = df_left.columns.intersection(df_right.columns)
        common_cols = common_cols.difference(on_keys)

        # 发出重复列警告
        if not common_cols.empty:
            warnings.warn(
                f"发现重复列: {list(common_cols)}. "
                f"{'保留最后一个DataFrame的值' if keep_last else '保留第一个DataFrame的值'}"
            )

        # 如果没有重复列，直接合并
        if common_cols.empty:
            return pd.merge(df_left, df_right, how=how, on=on_keys)

        # 使用默认的重命名策略处理重复列
        suffix = '_temp'
        rename_dict = {col: col + suffix for col in common_cols}
        df_left_renamed = df_left.rename(columns=rename_dict)

        # 合并DataFrame
        merged = pd.merge(df_left_renamed, df_right, how=how, on=on_keys)

        # 处理重复列
        for col in common_cols:
            temp_col = col + suffix
            if keep_last:
                # 保留最后一个DataFrame的值
                merged[col] = merged[col].combine_first(merged[temp_col])
            else:
                # 保留第一个DataFrame的值
                merged[col] = merged[temp_col].combine_first(merged[col])
            merged.drop(columns=[temp_col], inplace=True)

        return merged

    # 发出关于合并方式的警告
    if how in ['left', 'right']:
        warnings.warn(
            f"使用 how='{how}' 时，合并结果可能受DataFrame顺序影响。"
            "考虑使用 how='outer' 或 how='inner' 以获得更可预测的结果。"
        )

    # 使用reduce逐步合并所有DataFrame
    try:
        result = functools.reduce(merge_two_dfs, dfs)

        # 检查结果是否为空
        if result.empty:
            warnings.warn("合并后的DataFrame为空。请检查合并键和合并方式。")

        return result
    except Exception as e:
        raise ValueError(f"合并过程中发生错误: {str(e)}") from e

# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 将DataFrame的列索引按指定列表顺序排序，然后按指定列对数据进行排序
def df_sort(
        df: pd.DataFrame,
        cols: Optional[List[str]] = None,
        sort_by: Union[str, List[str]] = 'A',
        ascending: Union[bool, List[bool]] = True,
        na_position: str = 'last',
        keep_extra_cols: bool = True,
        inplace: bool = False) -> pd.DataFrame:
    """
    将DataFrame的列索引按指定列表顺序排序，然后按指定列对数据进行排序

    参数:
    df: 要排序的pandas DataFrame
    cols: 列顺序列表，如果为None，则保持原有列顺序
    sort_by: 用于行排序的列名或列名列表，默认为'A'
    ascending: 排序方向，True为升序，False为降序
                如果是列表，则对应每个排序列的排序方向
    na_position: 缺失值的位置，'first'或'last'，默认为'last'
    keep_extra_cols: 当为True时，不在cols列表中的列也包括在结果中；
                     当为False时，不在cols列表中的列不包括在结果中（被丢弃）
    inplace: 是否原地修改DataFrame，默认为False

    返回:
    排序后的DataFrame

    异常:
    ValueError: 当参数值无效时
    TypeError: 当参数类型不正确时
    """
    # 参数验证
    if not isinstance(df, pd.DataFrame):
        raise TypeError("df 必须是pandas DataFrame")

    if cols is not None and not isinstance(cols, list):
        raise TypeError("cols 必须是列表或None")

    if not isinstance(keep_extra_cols, bool):
        raise TypeError("keep_extra_cols 必须是布尔值")

    if na_position not in ['first', 'last']:
        raise ValueError("na_position 必须是 'first' 或 'last'")

    # 检查排序列是否存在
    if isinstance(sort_by, str):
        sort_columns = [sort_by]
    else:
        sort_columns = list(sort_by)

    missing_sort_cols = [col for col in sort_columns if col not in df.columns]
    if missing_sort_cols:
        raise ValueError(f"DataFrame中缺少排序列: {missing_sort_cols}")

    # 决定是否创建副本
    if not inplace:
        df = df.copy()

    # 处理列排序
    if cols is not None:
        # 检查cols中是否包含不存在的列
        extra_cols = [col for col in cols if col not in df.columns]
        if extra_cols:
            warnings.warn(f"cols中包含DataFrame中不存在的列: {extra_cols}")

        # 获取存在的列
        existing_cols = [col for col in cols if col in df.columns]

        # 确定最终列顺序
        if keep_extra_cols:
            # 保留不在cols中的列，放在指定列后面
            other_cols = [col for col in df.columns if col not in existing_cols]
            final_cols = existing_cols + other_cols
        else:
            # 只保留cols中存在的列，丢弃不在cols中的列
            final_cols = existing_cols

        # 重新排列列
        df = df[final_cols]

    # 按指定列排序
    df = df.sort_values(
        by=sort_columns,
        ascending=ascending,
        na_position=na_position
    )

    # 重置索引（可选，但通常排序后会重置索引）
    df.reset_index(drop=True, inplace=True)

    return df

# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 对DataFrame中指定列进行排名，返回形状相同的DataFrame
def df_rank_cols(
        df: pd.DataFrame,
        cols: List[str],
        method: Literal['min', 'max', 'average', 'first', 'dense'] = 'min',
        ascending: bool = True,
        na_option: Literal['keep', 'top', 'bottom'] = 'keep'    ) -> pd.DataFrame:
        """
        对DataFrame中指定列进行排名，返回形状相同的DataFrame

        参数:
        df (pd.DataFrame): 输入的DataFrame
        cols (list): 需要排名的列名列表
        method (str): 排名方法 ('min', 'max', 'average', 'first', 'dense')
        ascending (bool): 是否升序排列（True: 小值排名靠前，False: 大值排名靠前）
        na_option (str): NaN处理方式 ('keep', 'top', 'bottom')
            - 'keep': 保留原始NaN值（默认，推荐）
            - 'top': 将NaN视为最小值（排名1）
            - 'bottom': 将NaN视为最大值（排名最后）

        返回:
        pd.DataFrame: 指定列被替换为排名后的值，其他列保持不变

        为什么没有"视为0"的选项？
        --------------------------------------------------------
        1. 排名逻辑从1开始，0不是有效排名值
           例如：[1, 2, NaN] 的排名应为 [1, 2, ?]，不是 [1, 2, 0]

        2. pandas原生rank()函数不支持"0"选项
           pandas的rank()方法仅支持:
              na_option='keep' (默认)
              na_option='top'
              na_option='bottom'

        3. "视为0"会破坏排名逻辑
           - 排名表示"位置"（1=最高/最低，2=次高/次低...）
           - 0在排名中没有意义（排名从1开始）

        4. 正确做法：先排名，再处理NaN
           # 排名后将NaN替换为0（仅当需要时）
           result = df_rank_cols(df, cols)
           result = result.fillna(0)
        """
        df_ranked = df.copy()
        for col in cols:
            df_ranked[col] = df_ranked[col].rank(
                method=method,
                ascending=ascending,
                na_option=na_option
            )
        return df_ranked

# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 为 DataFrame 添加三列：行和、点积、点积排名（支持左右位置控制）
def df_add_rank(
        df: pd.DataFrame,
        lst: Optional[List[Union[int, float]]] = None,
        sum_col_name: str = "行和",
        dot_col_name: str = "点积",
        rank_col_name: str = "点积排名",
        direction: Literal['first', 'last'] = 'first',
        position: Literal['left', 'right'] = 'left'    ) -> pd.DataFrame:
    """
    为 DataFrame 添加三列：行和、点积、点积排名（支持左右位置控制）

    功能：
      - 从数值列中按方向（前/后）选 len(lst) 个列
      - 计算：行和、点积（加权和）、点积降序排名
      - 可选将三列置于最左或最右

    参数：
      df: 输入 DataFrame
      sum_col_name: 行和列名
      dot_col_name: 点积列名
      rank_col_name: 排名列名
      lst: 权重列表，长度决定参与计算列数（None → 全 NaN）
      direction: 'first'（前）或 'last'（后）
      position: 'left' 或 'right'

    返回：
      新增三列的 DataFrame（不修改原数据）
    """

    # ========== 1. 参数验证 ==========
    if not isinstance(df, pd.DataFrame):
        raise TypeError("df 必须是 pandas.DataFrame")

    if lst is not None and not isinstance(lst, list):
        raise TypeError("lst 必须是列表或 None")

    # ========== 2. 创建副本并提取数值列 ==========
    df = df.copy()
    numeric_df = df.select_dtypes(include=[np.number])

    if numeric_df.empty:
        # 直接创建三列 NaN 并返回
        return _add_nan_cols_and_reorder(df, [sum_col_name, dot_col_name, rank_col_name], position)

    numeric_columns = numeric_df.columns.tolist()
    n_numeric = len(numeric_columns)

    # ========== 3. 处理特殊 lst 情况 ==========
    if lst is None:
        return _add_nan_cols_and_reorder(df, [sum_col_name, dot_col_name, rank_col_name], position)

    if not lst:  # 空列表
        return _add_zero_cols_and_reorder(df, [sum_col_name, dot_col_name, rank_col_name], position)

    # ========== 4. 选择参与计算的列 ==========
    n_weights = len(lst)

    if n_weights > n_numeric:
        # 自动调整权重列表长度
        lst = lst[:n_numeric]
        selected_columns = numeric_columns
    else:
        selected_columns = (
            numeric_columns[:n_weights] if direction == 'first'
            else numeric_columns[-n_weights:]
        )

    # ========== 5. 向量化计算 ==========
    selected_data = df[selected_columns]
    weights_array = np.array(lst)

    # 一次性计算所有行
    row_sums = selected_data.sum(axis=1)
    dot_products = selected_data.dot(weights_array)

    # 使用更高效的排名方法
    dot_ranks = dot_products.rank(method='dense', ascending=False).astype(int)

    # ========== 6. 添加列并调整顺序 ==========
    df[sum_col_name] = row_sums
    df[dot_col_name] = dot_products
    df[rank_col_name] = dot_ranks
    return _move_cols(df, [sum_col_name, dot_col_name, rank_col_name], position)

# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
def _add_nan_cols_and_reorder(df: pd.DataFrame, col_names: List[str], position: str) -> pd.DataFrame:
    """快速添加 NaN 列并重排序"""
    for col in col_names:
        df[col] = np.nan
    return _move_cols(df, col_names, position)

# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
def _add_zero_cols_and_reorder(df: pd.DataFrame, col_names: List[str], position: str) -> pd.DataFrame:
    """快速添加零值列并重排序"""
    n_rows = len(df)
    df[col_names[0]] = 0.0  # 行总和
    df[col_names[1]] = 0.0  # 点积
    df[col_names[2]] = 1  # 排名（全部为1）
    return _move_cols(df, col_names, position)

# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
def _move_cols(df: pd.DataFrame, target_cols: List[str], pos: Literal['left', 'right']) -> pd.DataFrame:
    """将指定列移动到最左或最右，其余列保持原顺序"""
    existing_cols = [col for col in target_cols if col in df.columns]
    other_cols = [col for col in df.columns if col not in existing_cols]

    new_order = existing_cols + other_cols if pos == 'left' else other_cols + existing_cols
    return df[new_order]


# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■  封装为bytesIO/zip  ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 把一个 df 封装为一个二进进制的BytesIO对象。
def df_to_bytesIO(df):
    """
    将一个Pandas DataFrame对象封装为一个二进制的BytesIO对象。

    :param df: Pandas DataFrame对象
    :return: 一个包含Excel数据的二进制BytesIO对象
    """
    if not isinstance(df, pd.DataFrame):
        raise ValueError("df 必须是Pandas DataFrame对象")

    # 创建BytesIO对象
    bio_file = BytesIO()

    # 使用ExcelWriter将DataFrame写入BytesIO对象
    with pd.ExcelWriter(bio_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)  # index=False表示不包含行索引

    # 确保指针位于文件的开头
    bio_file.seek(0)
    return bio_file

# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 把一个 openpyxl 生成的 workbook 封装为一个二进进制的 BytesIO 对象。
def wb_to_bytesIO(wb):
    """
    把一个 openpyxl 生成的 workbook 对象封装为一个二进制的 BytesIO 对象。

    :param wb: 由 openpyxl 生成的 workbook 对象
    :return: 一个二进制的 BytesIO 对象
    """
    if not isinstance(wb, openpyxl.Workbook):
        raise ValueError("wb 必须是 openpyxl 生成的 Workbook 对象")

    bio_file = BytesIO()
    wb.save(bio_file)
    bio_file.seek(0)  # 确保指针位于文件的开头
    return bio_file

# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 多个DataFrame保存到内存中的ZIP文件,以提供给下载按钮
def dfs_to_zip(
        dfs_dic: Dict[str, pd.DataFrame],
        format: str = 'excel',
        empty_msg: str = '空值') -> BytesIO:
    """
    将多个DataFrame保存到内存中的ZIP文件

    Parameters:
    -----------
    dfs_dic : dict
        包含DataFrame的字典，键为文件名（不含扩展名）
    format : str, optional
        输出格式，支持'excel'（默认）或'csv'
    empty_msg : str, optional
        空DataFrame时显示的消息，默认为'空值'

    Returns:
    --------
    BytesIO
        包含ZIP文件内容的字节缓冲区

    Raises:
    -------
    ValueError
        如果指定的格式不被支持
    """
    # 验证格式参数
    if format not in ('excel', 'csv'):
        raise ValueError(f"不支持的格式: {format}. 支持 'excel' 或 'csv'")

    # 创建内存中的ZIP文件
    bio_zip = BytesIO()

    try:
        with zipfile.ZipFile(bio_zip, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for name, df in dfs_dic.items():
                # 检查DataFrame是否为空
                if df.empty:
                    df = pd.DataFrame({'提示': [empty_msg]})

                # 安全处理文件名
                safe_name = _sanitize_filename(str(name) if name is not None else 'data')

                # 根据格式处理数据
                if format == 'excel':
                    file_data = _df_to_excel(df, safe_name)
                    file_ext = 'xlsx'
                else:  # csv
                    file_data = _df_to_csv(df)
                    file_ext = 'csv'

                # 将文件数据写入ZIP
                zipf.writestr(f'{safe_name}.{file_ext}', file_data)

    except Exception as e:
        # 重新抛出异常，但先确保缓冲区被重置
        bio_zip.seek(0)
        bio_zip.truncate(0)
        raise e

    # 将指针重置到缓冲区开头
    bio_zip.seek(0)
    return bio_zip

# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 创建一个安全处理文件名的函数
def _sanitize_filename(
        name: str,
        max_length: int = 30) -> str:
    """
    安全处理文件名，移除非法字符并限制长度

    Parameters:
    -----------
    name : str
        原始文件名
    max_length : int, optional
        最大长度限制，默认为30

    Returns:
    --------
    str
        处理后的安全文件名
    """
    # 移除非ASCII字符和非法文件名字符
    safe_name = re.sub(r'[^\w\s-]', '', name).strip()

    # 替换空格为下划线
    safe_name = re.sub(r'[-\s]+', '_', safe_name)

    # 限制长度
    if len(safe_name) > max_length:
        safe_name = safe_name[:max_length]

    # 如果为空则使用默认名称
    if not safe_name:
        safe_name = 'data'

    return safe_name

# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 创建一个将DataFrame转换为Excel字节数据的函数
def _df_to_excel(
        df: pd.DataFrame,
        sheet_name: str) -> bytes:
    """
    将DataFrame转换为Excel字节数据

    Parameters:
    -----------
    df : pd.DataFrame
        要转换的DataFrame
    sheet_name : str
        Excel工作表名称

    Returns:
    --------
    bytes
        Excel文件的字节数据
    """
    excel_buffer = BytesIO()

    try:
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)  # Excel限制工作表名31字符
        return excel_buffer.getvalue()
    finally:
        excel_buffer.close()

# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
#　创建一个将DataFrame转换为CSV字节数据的函数
def _df_to_csv(
        df: pd.DataFrame) -> bytes:
    """
    将DataFrame转换为CSV字节数据

    Parameters:
    -----------
    df : pd.DataFrame
        要转换的DataFrame

    Returns:
    --------
    bytes
        CSV文件的字节数据（UTF-8编码）
    """
    csv_buffer = BytesIO()

    try:
        # 使用UTF-8编码确保中文正确显示
        df.to_csv(csv_buffer, index=False, encoding='utf-8-sig')
        return csv_buffer.getvalue()
    finally:
        csv_buffer.close()


# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■    字典相关的函数  ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 将字典的键值反转，并按照指定的顺序对反转后的键进行排序。
def dict_rev_sort(
        dic: Dict[Any, List[str]],
        sort_order: List[str],
        keep_all_keys: bool = False,
        default_value: Any = None    ) -> Dict[str, Any]:
    """
    将字典的键值反转，并按照指定的顺序对反转后的键进行排序。

    参数:
        dic: 原始字典，键为任意类型，值为字符串列表
        sort_order: 指定键的排序顺序列表
        keep_all_keys: 是否保留所有键（包括不在sort_order中的键）
        default_value: 对于不在sort_order中的键，使用的默认值（当keep_all_keys为False时使用）

    返回:
        反转并排序后的字典，键为原始字典值中的字符串，值为原始字典的键

    异常:
        ValueError: 当sort_order为空或dic为空时
        TypeError: 当参数类型不正确时
    """
    # 参数验证
    if not dic:
        raise ValueError("dic 不能为空")

    if not sort_order:
        raise ValueError("sort_order 不能为空")

    if not isinstance(dic, dict):
        raise TypeError("dic 必须是字典类型")

    if not isinstance(sort_order, list):
        raise TypeError("sort_order 必须是列表类型")

    # 反转字典
    reversed_dict = {}
    for key, value_list in dic.items():
        if not isinstance(value_list, list):
            raise TypeError(f"dic 的值必须是列表类型，但 {key} 的值是 {type(value_list)}")

        for item in value_list:
            reversed_dict[item] = key

    # 按照指定的顺序创建新字典
    sorted_dict = {}
    for item in sort_order:
        if item in reversed_dict:
            sorted_dict[item] = reversed_dict[item]
        elif not keep_all_keys and default_value is not None:
            sorted_dict[item] = default_value

    # 添加可能不在sort_order中的键（如果keep_all_keys为True）
    if keep_all_keys:
        for item, value in reversed_dict.items():
            if item not in sorted_dict:
                sorted_dict[item] = value

    return sorted_dict