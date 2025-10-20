from io import BytesIO
import numpy as np
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

'''
# 2024.7.8更新,对多个函数进行优化
'''

xk_dic = {"语文": 1, "数学": 2, "英语": 3, "物理": 4, "化学": 5, "生物": 6, "政治": 7, "历史": 8, "地理": 9}
# 优化了上面的代码，提高了代码的逻辑和可读性。
# subject_lst = ["语文", "数学", "英语", "物理", "化学", "生物", "政治", "历史", "地理"]
# subject_dic= {val: idx+1 for idx, val in enumerate(subject_lst)}

class ana_df():
    def __init__(self, df, xk_dic):
        # 确定df表的标准学科名的列表:xkm.
        xkm = list(set(xk_dic.keys()) & set(df.columns))
        xkm.sort(key=lambda x: xk_dic[x])
        # 增加总分列
        df["总分"] = df.loc[:, xkm].sum(axis=1)
        # 增加班次列
        df["班次"] = df.groupby("班级")["总分"].rank( ascending=False, method="min")
        # 增加级次列
        df["级次"] = df["总分"].rank(axis=0, ascending=False, method="min")

        self.__df = df        # 原始数据表
        self.__xkm = xkm      # 学科名的列表

    def get_all(self):
        """
        返回ana_df表的全部信息,包括[班级,学号,姓名,所有学科,总分,班次,级次]所有列.
        :return:pd的DataFrame数据表类型.
        """
        return self.__df  # 返回表的全部信息

    def get_df(self, bas=None, xk=1, count=None, banc=40):
        """
        返回ana_df数据表的部分列.
        :param bas: 基础信息的列标,取值范围为:["班级", "学号", "姓名"],默认值为全选
        :param xk: 基础信息的列标,取值范围为:["语文", "数学", "英语", "物理", "化学", "生物", "政治", "历史", "地理" ],默认值1,表示返回全部学科数据.
        :param count: 计算信息的列表,取值范围为:["总分", "班次", "级次"],默认值为空,默认不返回列数据.
        :param banc: 整型数据.默认值为40,表示计算班级前40名同学成绩
        :return: pd的DataFrame数据表类型.
        """

        df = self.__df
        xkm = self.__xkm
        if banc != 0:
            df = df[df["班次"] <= banc]
        if bas is None:
            bas = ["班级", "学号", "姓名"]
        if count is None:
            count = []
        if xk == 0 or None:
            return df.loc[:, bas + count]
        elif xk == 1:
            return df.loc[:, bas + xkm + count]

    def get_mc(self, bas=None, banc=40,tup=1):
        """
        返回一个学科与总分的名次表.
        :param tup: 当tup=0时，每个单元格是一个整型名次数据。当tup=1时，每个单元格是学科，名次封装的一个元组:(学科名次,总分名次),总分名次为一个整型数据.
        :param bas: 学科名称的列表.取值范围为:["班级", "学号", "姓名"],默认值为:bas = ["班级", "姓名"]
        :param banc: 整型数据.默认值为40,表示选择40名同学
        :return: pd的DataFrame数据表类型.
        """
        df = self.__df
        if banc != 0:
            df = df[df["班次"] <= banc]
        if bas is None:
            bas = ["班级", "学号","姓名"]

        df_mc = df.loc[:, self.__xkm + ['总分']].rank(axis=0, method="min", ascending=False)
        if tup == 1:
            for i in range(len(self.__xkm)):
                df_mc.iloc[:, i] = df_mc.iloc[:, [i, len(self.__xkm)]].apply(tuple, axis=1)  # 将多列的名次数据与总分名次列合并为多个元组列.
        return pd.concat([df[bas], df_mc], axis=1)

    def get_xkm(self):
        '''
        获取实有的学科列标签.
        :return:列表类型.学科的列标签的列表.
        '''
        return self.__xkm

def Fun_sdb(sr, xk, zf):
    '''
    双达标函数:统计一个SR数据中,学科小于等于xk,且总分小于等于zf的数据的个数
    :param sr: pd的SR数据类型,一维数据,每个数据是一个元组(a,b).
    :param xk: 学科名次,确定学科排名的最大值.
    :param zf: 总分名次,确定总分排名的最大值.
    :return: int类型,整型数据,即在sr中,满足学科排名小于xk,总分排名小于zf的个数.
    '''
    s = 0
    for i in sr:
        if (i[0] <= xk) & (i[1] <= zf):
            s += 1
    return s

def GFuns_fsd(ls):
    '''
    返回一个不同名的函数组,每个函数统计落在区间的数据的个数.统计时,前闭后开.
    :param ls: 列表类型.是一个成绩阈值的列表,如:[36, 48, 60, 72, 78, 84, 90, 96, 102, 108, 114, 120].
    :return: 一个不同名的函数组,每个函数统计落在某个分数段中的数据的个数.
        其中某个函数举例为:F_1(arr),形参arr为列表数据类型.
    '''
    funcs = []
    for i in range(len(ls)):
        if i == len(ls) - 1:
            exec(f"def F_{i + 1}(arr,p={i}):return sum(arr >= {ls}[p])")
        else:
            exec(f"def F_{i + 1}(arr,p={i}):return sum((arr >= {ls}[p]) & (arr < {ls}[p + 1]))")
        funcs.append(locals()[f"F_{i + 1}"])
    return funcs

def GFunS_mcd(ls,lj=0):
    '''
    返回一个不同名的函数组,每个函数统计落在区间的数据的个数.统计时,(重点)前开后闭.
    :param ls: 列数类型.是一个成绩阈值的列表,如:[0, 10, 50, 100, 150, 200, 250, 300, 350, 400]
    :param lj: 数据是否累计,默认不累计
    :return: 一个不同名的函数组,每个函数统计落在某个名次段中的数据的个数.
            其中某个函数举例为:MC(xx-xx](arr),形参arr为列表数据类型.
    '''
    # 确保 ls 是一个列表或元组
    if not isinstance(ls, (list, tuple)):
        raise ValueError("ls必须为列表或元组")

    def make_func(lower, upper):
        def func(arr):
            arr = np.array(arr)
            # 使用 numpy 的逻辑与和比较来找到区间内的元素
            count = np.sum((arr > lower) & (arr <= upper))
            return count
        return func

    funcs = []
    # 处理所有区间，除了最后一个区间是半开的（使用 float('inf') 作为上限）
    for i in range(len(ls) - 1):
        if lj == 0:
            lower = ls[i]
        else:
            lower = ls[0]
        upper = ls[i + 1]
        func = make_func(lower, upper)
        func.__name__ = f'MC_{lower}-{upper}-{lj}'  # 设置函数名，包含区间信息
        funcs.append(func)

    # 处理最后一个区间（包含所有大于最后一个阈值的元素）
    last_func = make_func(ls[-1], float('inf'))
    last_func.__name__ = f'MC_{ls[-1]}-up-{lj}'
    funcs.append(last_func)
    # 返回函数
    return funcs

def GFunS_mcd_tup(lst, max_score):
    '''
    返回一个函数组，每个函数统计落在指定分数区间内的数据个数（前开后闭）。

    :param lst: 列表类型，包含分数阈值，必须是升序排列。
    :param max_score: 整数或浮点数，用于定义最高分数边界（不包含）。
    :return: 一个函数列表，每个函数统计落在某个分数段中的数据个数。
    '''
    if not isinstance(lst, (list, tuple)) or not all(isinstance(x, (int, float)) for x in lst):
        raise ValueError("thresholds必须为升序排列的整数或浮点数列表或元组")
    if not isinstance(max_score, (int, float)):
        raise ValueError("max_score必须为整数或浮点数")

    def make_counter(lower, upper):
        def counter(data):
            count = 0
            for score, xmc in data:  # 假设data中的每个元素都是一个(score, other_info)的元组
                if lower < score <= upper  and xmc<=max_score:
                    count += 1
            return count
        return counter

    counters = []
    for i in range(len(lst)):
        lower = lst[i]
        upper = lst[i + 1] if i < len(lst) - 1 else max_score  # 使用max_score作为上界（不包含）
        counter = make_counter(lower, upper)
        counter.__name__ = f'MC_{lower}-{upper if upper != max_score else "up"}'
        counters.append(counter)

        # 处理最后一个区间（包含所有大于或等于最后一个阈值且小于max_score的元素）

    def last_counter(data):
        count = 0
        for score, _ in data:
            if score >= lst[-1] and score < max_score:  # 使用thresholds[-1]作为下界
                count += 1
        return count

    counters.append(last_counter)
    return counters

def GFuns_lv(ls):
    '''
    返回一个不同名的函数组,如计算:及格人数,及格率,优秀人数,优秀率.
    :param ls: 列数类型.是一个成绩阈值的列表,如:[60,80,100].用于计算及格率,优秀率
    :return: 一个不同名的函数组,每个函数统计落及格率,优秀率.
    '''
    thresholds=ls
    funcs = []
    # 辅助函数，用于生成人数和比率计算的函数
    def make_funcA(lower, upper, label):
        def func(arr):
            arr = np.array(arr)
            count = np.sum((arr >= lower) & (arr <= upper))
            return count
        func.__name__ = f'F_{label}_count'  # 设置函数名（可选，用于调试）
        return func

    def make_funcB(lower, upper, label):
        def func(arr):
            arr = np.array(arr)
            count = np.sum((arr >= lower) & (arr <= upper))
            ratio = count / len(arr) if len(arr) > 0 else 0
            return ratio
        func.__name__ = f'F_{label}_ratio'  # 设置函数名（可选，用于调试）
        return func
    for i in range(len(thresholds) - 1):
        lower = thresholds[i]
        upper = thresholds[-1]
        funcs.append(make_funcA(lower, upper, f'{lower}-{upper}'))
        funcs.append(make_funcB(lower, upper, f'{lower}-{upper}'))
    return funcs

def Fun_jf(arr, qz):
    '''
    对两个等长列表对应值的积求和, 定义积分函数
    :param arr: 数组,列表类型.
    :param qz: 权重,列表类型.
    :return: 浮点因型.对两个列表的对应值求积,把所有的积求和.
    '''
    return sum(m * n for m, n in zip(arr, qz))

def retain_worksheet(wb, sheet_name):
    """
    只保留指定名称的工作表，并删除工作簿中的其他所有工作表。

    :param wb: openpyxl工作簿对象
    :param sheet_name: 要保留的工作表名称
    :return: 修改后的(工作簿对象,保留的工作表对象)
    """
    # 遍历工作簿中的所有工作表
    for sheet in list(wb.worksheets):  # 注意：使用list()避免在迭代时修改集合
        if sheet.title != sheet_name:
            wb.remove(sheet)  # 删除非指定名称的工作表

    # 返回工作簿和保留的工作表对象
    # 注意：这里直接返回wb和wb[sheet_name]即可，因为wb[sheet_name]就是保留的工作表对象
    return wb, wb[sheet_name]

def ls_to_ws(ws, cells, ls):
    """
    把列表的值填充到openpyxl创建的ws表的指定单元格中。

    :param ws: 由openpyxl创建的Worksheet对象。
    :param cells: 一个坐标列表，如[(2, 2), (2, 4), (3, 3), (3, 11)]，表示要填充的单元格位置。
    :param ls: 一个Python列表，包含要填充的值。
    :return: 无返回值

    注意：确保列表`ls`的长度与`cells`的长度相匹配，否则将引发异常。
    """
    # 检查ws是否为Worksheet对象
    if not hasattr(ws, 'cell'):
        raise ValueError("ws必须是一个openpyxl的Worksheet对象")

        # 检查cells和ls的长度是否相等
    if len(cells) != len(ls):
        raise ValueError("cells和ls的长度必须相等")

        # 遍历cells和ls，将值填充到对应的单元格中
    for cell, val in zip(cells, ls):
        # 确保cell是一个包含两个元素的元组或列表，分别代表行和列
        if not (isinstance(cell, tuple) or isinstance(cell, list)) or len(cell) != 2:
            raise ValueError("cells中的每个元素都必须是一个包含两个元素的元组或列表")

            # 填充单元格
        ws.cell(row=cell[0], column=cell[1], value=val)

def sr_to_ws(ws, cells, sr):
    """
    把Pandas Series的值填充到openpyxl创建的ws表的指定单元格中。

    :param ws: 由openpyxl创建的Worksheet对象。
    :param cells: 一个坐标列表，如[(2, 2), (2, 4), (3, 3), (3, 11)]，表示要填充的单元格位置。
    :param sr: 一个Pandas Series对象，包含要填充的值。
    :return: 无返回值
    """
    # 确保sr是一个Pandas Series
    if not isinstance(sr, pd.Series):
        raise ValueError("sr必须是一个Pandas Series对象")

        # 确保sr的长度与cells的长度相匹配
    if len(sr) != len(cells):
        raise ValueError("sr的长度必须与cells列表的长度相匹配")

        # 处理sr中的值，将NaN转换为None（或其他占位符）
    values = sr.fillna('None').tolist()  # 也可以使用.fillna('')来替换为空字符串

    # 遍历cells和values，将值填充到对应的单元格中
    for cell, val in zip(cells, values):
        ws.cell(row=cell[0], column=cell[1], value=val)

def df_to_ws(ws, cells, df):
    """
    把单行DataFrame的值填充到ws表(openpyxl创建)的指定单元格中。

    :param ws: 由openpyxl创建的Worksheet对象。
    :param cells: 一个坐标列表，如[(2, 2), (2, 4), (3, 3), (3, 11)]，表示要填充的单元格位置。
    :param df: 一个只有一行的DataFrame，包含要填充的值。
    :return: 无返回值
    """
    # 确保df确实只有一行
    if df.shape[0] != 1:
        raise ValueError("df必须只有一行")

    # 获取df中唯一一行的值，转换为列表
    values = df.iloc[0].tolist()

    # 遍历cells和values，将值填充到对应的单元格中
    for cell, val in zip(cells, values):
        ws.cell(row=cell[0], column=cell[1], value=val)

def dfs_to_ws(ws, row, col, dfs, rg=0, cg=0,hd=False):
    """
    将多个Pandas DataFrame逐一注入到一个由openpyxl创建的工作表中。

    :param ws: 由openpyxl创建的工作表。
    :param row: 第一个DataFrame的行坐标（起始行）。
    :param col: 第一个DataFrame的列坐标（起始列）。
    :param dfs: DataFrame所组成的列表。
    :param rg: DataFrame之间的行间距。
    :param cg: DataFrame之间的列间距（通常保持为0，因为列宽是固定的）。
    :return: None. 只有注入操作，返回值为空。
    """
    for df_e in dfs:
        # 遍历DataFrame的每一行
        for r, row_data in enumerate(dataframe_to_rows(df_e, index=False, header=hd), start=row):
            # 将每行数据写入到工作表的对应位置
            for c, value in enumerate(row_data, start=col):
                if pd.isnull(value):
                    value = None  # 或选择转换为特定值
                ws.cell(row=r, column=c, value=value)

                # 更新行位置为下一个DataFrame的起始位置
        row += rg

def df_split(df):
    """
    将一个具有多级列索引的DataFrame按列索引的第0级拆分成多个单独的DataFrame。
    参数: df (pandas.DataFrame): 输入的DataFrame，它应该具有多级列索引。
    返回: list of pandas.DataFrame: 一个包含多个DataFrame的列表，每个DataFrame包含具有相同第0级列索引的列。
    注意:
        - 如果输入的DataFrame没有多级列索引，函数将抛出ValueError异常。
        - 如果第0级列索引中有重复项，则每个重复项对应的列都将被包含在同一个拆分后的DataFrame中。
    """
    # 确保df确实具有多级列索引
    if not isinstance(df.columns, pd.MultiIndex):
        raise ValueError("df的列索引必须是MultiIndex")

    # 获取所有唯一的第0级列标签
    unique_level_0_labels = df.columns.get_level_values(0).unique()

    # 初始化一个空列表来存储拆分后的DataFrame
    split_dfs = []

    # 遍历每个唯一的第0级标签
    for label in unique_level_0_labels:
        # 选择所有具有该第0级标签的列
        mask = df.columns.get_level_values(0) == label
        selected_columns = df.columns[mask]

        # 使用这些列来创建一个新的DataFrame
        temp_df = df[selected_columns]

        # 将新的DataFrame添加到列表中
        split_dfs.append(temp_df)

        # 返回包含所有拆分后DataFrame的列表
    return split_dfs

def df_divide(df, col):
    """
    把一个df表,按col的列标签,拆分为多个df表.
    这是一个旧的函数,其功能由更新的函数split_df_by_column代替,

    :param df:  df表类型
    :param col: 列标签
    :return: 多个df表组成的列表.
    """
    ls = []
    for name, g in df.groupby(col):
        ls.append(g)
    return ls

def split_df_by_column(df, col):
    """
    按指定的列标签将DataFrame拆分为多个子DataFrame，并将它们存储在字典中返回。
    :param df: DataFrame对象
    :param col: 用于分组的列标签
    :return: 一个字典，键为分组名，值为对应的子DataFrame
    """
    if not isinstance(df, pd.DataFrame):
        raise ValueError("df必须是一个DataFrame对象")
    if col not in df.columns:
        raise ValueError(f"列标签'{col}'不在DataFrame中")

    return {name: group for name, group in df.groupby(col)}

def workbook_to_bytesIO(wb):
    '''
    把一个openpyxl生成的workbook对象,封装为一个二进制的BytesIO对象
    :param wb: 由openpyxl生成的workbook对象
    :return: 一个二进制的BytesIO对象
    '''
    BIO_file = BytesIO()
    wb.save(BIO_file)
    return BIO_file