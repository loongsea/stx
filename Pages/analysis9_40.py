from longsea import al
from openpyxl import load_workbook
import streamlit as st
import pandas as pd
import time
import os
from longsea import al2
from longsea.al2 import wb_to_bytesIO, df_sort, validate_dataframe

txt_A = '''
### 说明:
*  九年级2024算法.
*  al2.2025.9.26.
'''
txt_B = '''
### 版本更新:
* ###### 2022.10.1,开始学习numpy、pandas。
* ###### 2023.2.1,完成程序设计，必须为9个学科。
* ###### 2023.4.5,用streamlit制作了一个web外壳。
* ###### 2023.4.30,第一次重写代码，自适应若干个学科。
* ###### 2023.5.7-5.20,添加滑块,调用模板.
* ###### 2023.7.29,第二次重写代码,创建als模块.
* ###### 2024.7.7-10.23,简化一些代码.对九年级采用新分析方法。
* ###### 2024.11.14-23,把als重构为al。
* ###### 2025.9.21-928,把al重构为al2.
'''
long_text = '''

# =============================双达标报表================================
# 双达标的第一条件：学科名次段阈值列表.
mcd_thresh_ls=[0,200,260,300]
# 学科名次段阈값列表对应积分.
mcd_thresh_score = [10, 9, 2, 1, 0]       # 每个分数段的积分值
# 双达标的第二条件：最大校次.
max_total_rank = 260

# =============================两率一平报表===============================
# 学科方案字典
  dic_lv = {120: ["语文", "数学", "英语"],
            70: ["物理", "政治"],
            50: ["化学", "生物", "历史", "地理"]}
# 班级参评人数
  滑动条控制:max_class_rank。

# =============================班级分析报表===============================
# 班级名次段阈값列表
thresh_cls = [0, 10, 50, 100, 150, 200, 240, 300, 350, 400]      # 各个名次段
# 班级名次段阈값列表对应积分.
thresh_cls_score = [10, 10, 10, 10, 8, 6, 1, 1, 1, 1]          # 各个名次段的积分
# 班级参评人数由滑动条控置:max_cls_rank。

'''

# 模板路径信息
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# ===============================模板xlsx文件===========================================================
# 模板文件
MB_file = "../longsea/模板2024.xlsx"
# 获取当前脚本文件的绝对路径目录，确保无论从哪个工作目录运行程序都能正确找到文件
script_dir = os.path.dirname(os.path.abspath(__file__))
# 将脚本目录与模板文件名组合成完整路径，确保跨平台兼容性
mb_file_path = os.path.join(script_dir, MB_file)
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■


# 设置页面信息
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
st.set_page_config(page_title="成绩分析2025_al2",layout="centered", page_icon=":bar_chart:",initial_sidebar_state="expanded",)
# 显示主标题
st.markdown(''' ### 九年级分析2025-al2 ''')
st.write("***")
# 添加侧边栏说明文本
st.sidebar.write(txt_A)



# 设置交互控件
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 创建一个上传文件的按钮
uploaded_file = st.file_uploader("上传XLSX文件", type=["xlsx"])

# 提示上传文件
if not uploaded_file:
    st.text_area(label='设置说明',value=long_text,height=400)
    exit()
elif uploaded_file:
    col_A, col_B = st.columns(2)
    with col_A:
        dic_df = pd.read_excel(uploaded_file, engine='openpyxl', sheet_name=None)       # 读取所有工作表。
        selected_sheet_name = st.selectbox("选择工作表", list(dic_df.keys()))
        df = dic_df[selected_sheet_name]
        # 添加一个滑动条,用于选择统计学科成绩时,计算的班级学生数.
        max_class_rank = st.slider('两率一平参评人数', min_value=1, max_value=60, value=45)
    with  col_B:
        try:
            dic_mb_df = pd.read_excel(mb_file_path, engine='openpyxl', sheet_name=None)
            sht_MB_name = st.selectbox("选择模板", list(dic_mb_df.keys()))
            mb_df = dic_mb_df[sht_MB_name]
        except FileNotFoundError:
            st.error("找不到 test.xlsx 文件，请确保文件存在于当前目录中")
        except Exception as e:
            st.error(f"读取文件时发生错误: {str(e)}")
        # 添加一个滑动条,用于选择统计学科成绩时,计算的班级学生数.
        max_cls_rank = st.slider('班级分析参评人数', min_value=1, max_value=60, value=50)
    # if st.checkbox("全部显示"):
    #     st.dataframe(df,use_container_width= True)
    # else:
    #     st.dataframe(df.head(2),use_container_width= True)
    if st.checkbox("检查/显示数据"):
        vdf = validate_dataframe( df, required_columns=['班级', '姓名', '学号'], allow_extra_columns=False)
        st.dataframe(vdf,use_container_width= True)
        adf = al2.Andf(df)
    else:
        st.dataframe(df.head(2),use_container_width= True)
        adf = al2.Andf(df)

start = time.time()
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■   未启用报表(占位)  ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■





# ■■■■■■■■■■■■■■■■■■■■■■■■■■■       获取双达标报表(sdb_dfs)        ■■■■■■■■■■■■■■■■■■■■■■■■■■■
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 双达标的第一条件：学科名次段阈值列表.
mcd_thresh_ls=[0,180,240,300]
# 学科名次段阈값列表对应积分.
mcd_thresh_score = [10, 9, 2, 1, 0]       # 每个分数段的积分值
# 双达标的第二条件：最大校次.
max_total_rank = 240

# -----------------------------------------------------------------------------------------

sdb_dfs = adf.get_sdb(
    thresh = mcd_thresh_ls,
    thresh_score = mcd_thresh_score,
    max_total_rank = max_total_rank )


# ■■■■■■■■■■■■■■■■■■■■■■■■■■■       获取两率一平报表(lv_dfs)        ■■■■■■■■■■■■■■■■■■■■■■■■■■■
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 学科方案字典
dic_lv = {120: ["语文", "数学", "英语"],
          70: ["物理", "政治"],
          50: ["化学", "生物", "历史", "地理"]}
# 班级参评人数由滑动条控制:max_class_rank。
# -----------------------------------------------------------------------------------------

lv_dfs = adf.get_lv(
    dic_total_sbj = dic_lv,
    thresh =[0.6,0.8],
    include_count_valid=-1,
    max_class_rank=max_class_rank )


# ■■■■■■■■■■■■■■■■■■■■■■■■■■■         获取班级报表(bj_dfs)         ■■■■■■■■■■■■■■■■■■■■■■■■■■■
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 班级名次段阈값列表
thresh_cls = [0, 10, 50, 100, 150, 200, 240, 300, 350, 400]      # 各个名次段
# 班级名次段阈값列表对应积分.
thresh_cls_score = [10, 10, 10, 10, 8, 6, 1, 1, 1, 1]          # 各个名次段的积分
# 班级参评人数由滑动条控置:max_cls_rank。
# -----------------------------------------------------------------------------------------

bj_dfs = adf.get_cls(thresh=thresh_cls,
                     thresh_score=thresh_cls_score,
                     max_class_rank=max_cls_rank)

bj1_dfs = adf.get_cls(thresh=thresh_cls,
                      max_class_rank=max_cls_rank,
                      cumu=1 )

# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# B:选择模板文件ws,将分析报表注入工作表ws中.
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 打开模板文件。
wb = load_workbook(mb_file_path)  # 模板文件.
# 清洗工作簿wb,保存工作簿wb与模板工作表ws,
wb,ws = al.trim_wb(wb,sht_MB_name)

# 将双达标报表注入ws表中.
al2.dfs_to_ws(ws,5,4,sdb_dfs.values(),48,0,False,idx=False)
# 将两率一平报表注入ws表中.
al2.dfs_to_ws(ws,5,12,lv_dfs.values(),48,0,False,idx=False)
# 将班级报表注入ws表中.
al2.dfs_to_ws(ws,357,4,bj_dfs,16,0,False,idx=False)
al2.dfs_to_ws(ws,399,7,bj1_dfs,16,0,False,idx=False)

# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# C:选择模板,创建工作表,将成绩表、名次表注入工作薄中.
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 在模板文件中新建名次表ws1.
ws1 = wb.create_sheet("名次表",1)
# 获取名次表,并将数据注入ws1表中。
df_MC0 = adf.get_mc(combine_ranks=0)
df_MC0 = df_sort(df_MC0,cols=['班级','学号','姓名',"语文","数学","英语"],sort_by = ["班级","级次"])
al2.dfs_to_ws(ws1,1,1,df_MC0,hd=True)

# ==============================================================================================
# 在模板文件中新建成绩表ws2.
ws2 = wb.create_sheet("成绩表",1)
# 获取名次表,并将数据注入ws1表中。
df_all = adf.get_all()
df_all = df_sort(df_all,cols=['班级','学号','姓名',"语文","数学","英语"],sort_by = ["班级","级次"])
al2.dfs_to_ws(ws2,1,1,df_all,hd=True)


# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# C:添加一个下载按钮
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 添加时间信息。
st.success(f"运算己经完成，共用时：{round(time.time() - start, 2)}秒。")
# 创建下载按钮,以便下载此工作簿
st.download_button(
    label = '下载分析结果',
    data = wb_to_bytesIO(wb),
    file_name = os.path.splitext(uploaded_file.name)[0] + "_" + sht_MB_name + "_报表" + ".xlsx",
    mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')






