from longsea import al
from openpyxl import load_workbook
import streamlit as st
import pandas as pd
import time
import os
from longsea.al import wb_to_bytesIO

txt_A = '''
### 说明:
*  分析九年级的成绩.
*  采用2024年新算法.
*  自动生成分析报表.
*  重构als模块为al。
'''
txt_B = '''
### 版本更新:
* ###### 2022.10.1,开始学习numpy、pandas。
* ###### 2023.2.1,完成程序设计，必须为9个学科。
* ###### 2023.4.5,用streamlit制作了一个web外壳。
* ###### 2023.4.30,第一次重写代码，自适应若干个学科，不需要必须9个学科了。
* ###### 2023.5.7,添加滑块,确定学科分析的每班参评人数.
* ###### 2023.5.10,对学科名进行排序.
* ###### 2023.5.18,对分析结果调用模板.班级分析全班参与.
* ###### 2023.5.20,自动调用级段模板填充数据.班级分析人数可选.
* ###### 2023.7.29,第二次重写代码,采用了模块化/对像化/函数化技术.
* ###### 2024.7.7,简化一些代码.
* ###### 2024.10.23,对九年级采用名次段的新分析方法.
* ###### 2024.11.14-23,对al进行重构。
'''
long_text = '''
# 设置所有学科的名称列表。
xk_dic = ["语文", "数学", "英语", "物理", "化学", "生物", "政治", "历史", "地理"]

# 分数段报表：学科名次段阈值列表.
dic_fsd = {(0,72,96,100,120): ["语文", "数学", "英语"], (0,42,56,60,70): ["物理", "政治"], (0,30,40,45,50): ["化学","生物", "历史", "地理"]}
# 分数段报表：学科名次段阈值对应积分列表.
fsd_thresh_score=[5,4,3,1]

# 双达标报表：学科名次段阈值列表.
mcd_thresh_ls=[0,200,260,300]
# 双达标报表：学科名次段阈值列表对应积分.
mcd_thresh_score = [10, 9, 2, 1, 0]       # 每个分数段的积分值
# 双达标报表：全校最大校次.
max_total_rank=260
# 双达标报表：班级参评人数由滑动条控制:max_class_rank。

# 两率一平报表:学科方案字典
dic_lv = {120: ["语文", "数学", "英语"], 70: ["物理", "政治"], 50: ["化学", "生物", "历史", "地理"]}
# 两率一平报表:班级参评人数由滑动条控制:max_class_rank。

# 班级分析报表:班级名次段阈值列表
thresh_cls = [0, 10, 50, 100, 150, 200, 260, 300, 350, 400]      # 各个名次段
# 班级分析报表:班级名次段阈值列表对应积分.
thresh_cls_score = [10, 10, 10, 10, 8, 6, 2, 1.5, 1, 1]          # 各个名次段的积分
# 班级分析报表:班级参评人数由滑动条控置:max_cls_rank。

# 成绩报表的xlsx文件.
MB_file = "c:/loongsea/模板2024.xlsx"
'''


# ----------------------------------------------------------------------------------------------------------------------
# 设置参数信息
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■

# 设置所有学科的名称列表。
xk_dic = ["语文", "数学", "英语", "物理", "化学", "生物", "政治", "历史", "地理"]

# 分数段报表：学科名次段阈值列表.
dic_fsd = {(0,72,96,100,120): ["语文", "数学", "英语"],
           (0,42,56,60,70): ["物理", "政治"],
           (0,30,40,45,50): ["化学","生物", "历史", "地理"]}
# 分数段报表：学科名次段阈值对应积分列表.
fsd_thresh_score=[5,4,3,1]

# 双达标报表：学科名次段阈值列表.
mcd_thresh_ls=[0,200,260,300]
# 双达标报表：学科名次段阈值列表对应积分.
mcd_thresh_score = [10, 9, 2, 1, 0]       # 每个分数段的积分值
# 双达标报表：全校最大校次.
max_total_rank=260
# 双达标报表：班级参评人数由滑动条控制:max_class_rank。

# 两率一平报表:学科方案字典
dic_lv = {120: ["语文", "数学", "英语"],
          70: ["物理", "政治"],
          50: ["化学", "生物", "历史", "地理"]}
# 两率一平报表:班级参评人数由滑动条控制:max_class_rank。

# 班级分析报表:班级名次段阈值列表
thresh_cls = [0, 10, 50, 100, 150, 200, 260, 300, 350, 400]      # 各个名次段
# 班级分析报表:班级名次段阈值列表对应积分.
thresh_cls_score = [10, 10, 10, 10, 8, 6, 2, 1.5, 1, 1]          # 各个名次段的积分
# 班级分析报表:班级参评人数由滑动条控置:max_cls_rank。

# 成绩报表的xlsx文件.
# MB_file = "c:/loongsea/模板2024.xlsx"
MB_file = "c:/loongsea/模板2024.xlsx"


# ----------------------------------------------------------------------------------------------------------------------
# 设置页面信息
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 设置网页显示信息
st.set_page_config(page_title="成绩分析2024_LOONGSEA",layout="centered", page_icon=":bar_chart:",initial_sidebar_state="expanded",)
# 显示主标题
st.markdown(''' ### 成绩分析程序2024-重构al ''')
st.write("***")
# 添加侧边栏说明文本
st.sidebar.write(txt_A)

# 创建一个上传文件的按钮
uploaded_file = st.file_uploader("上传XLSX文件", type=["xlsx"])
if not uploaded_file:
    st.text_area(label='设置说明', value=long_text, height=480, help=txt_B)
    exit()
elif uploaded_file:
    # 添加两个分列.
    col_A, col_B = st.columns(2)
    # 左分列
    with col_A:
        # 使用Pandas读取己上传的Excel文件
        df = pd.read_excel(uploaded_file, engine='openpyxl', sheet_name=None)      # 读取所有工作表。
        # 创建一个选择框，返回列表中的一个选中的工作表。df.keys(),df的所有工作表名列表
        selected_sheet = st.selectbox("选择工作表", list(df.keys()))
        # 将所选工作表的数据返回df表。
        df = df[selected_sheet]                          # =====================================>>df表为所选工作表学生成绩表
        # 添加一个滑动条,用于选择统计学科成绩时,计算的班级学生数.
        max_class_rank = st.slider('学科分析参评人数', min_value=1, max_value=60, value=45)
    # 右分列:
    with col_B:
        # 读取模板文件
        df_MB = pd.read_excel(MB_file, engine='openpyxl', sheet_name=None)
        # 创建选择框，返回其中选中的工作表。
        sht_MB_name = st.selectbox("选择报表模板", list(df_MB.keys()))
        # 添加一个滑动条,用于选择统计学科成绩时,计算的班级学生数.
        max_cls_rank = st.slider('班级分析参评人数', min_value=1, max_value=60, value=50)

# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# A:读取文件并定义为ana_df对象.获取分析报表.
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■

start = time.time()

#  生成为ana_df对象df.
if df.empty:
    exit()
df = al.Andf(df)

# 获取双达标分析表(sdb_df)
sdb_dfs = df.get_sdb(thresh=mcd_thresh_ls,thresh_score=mcd_thresh_score,max_class_rank=max_class_rank,max_total_rank=max_total_rank)

# 获取两率一平表(df_lv)
lv_dfs = df.get_lv(dic_val_sbj=dic_lv,calcu=0,max_class_rank=max_class_rank)

# 获取班级分析报表(df_cls)
df_cls = df.get_cls(thresh=thresh_cls,thresh_score=thresh_cls_score,max_class_rank=max_cls_rank)

# 获取分数段报表(fsd_dfs)
# fsd_dfs = df.get_fsd(dic_val_sbj=dic_fsd,thresh_score=fsd_thresh_score,max_class_rank=max_class_rank)
# for dff in fsd_dfs:
#     st.dataframe(dff,width=800)

# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# B:选择模板文件ws,将分析报表注入工作表ws中.
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 打开模板文件。
wb = load_workbook(MB_file)  # 模板文件.
# 清洗工作簿wb,保存工作簿wb与模板工作表ws,
wb,ws = al.trim_wb(wb,sht_MB_name)

# 将班级报表注入ws表中。
al.dfs_to_ws(ws,133,4,df_cls)
# 将双达标报表注入ws表中.
al.dfs_to_ws(ws,5,4,sdb_dfs,16,0,hd=False)
# 将率平报表注入ws表中。
al.dfs_to_ws(ws,5,12,lv_dfs,16,0,hd=False)

# 在模板文件中新建名次表ws1.
ws1 = wb.create_sheet("名次表",1)
# 获取名次表,并将数据注入ws1表中。
df_MC0 = df.get_mc(max_class_rank=max_cls_rank,combine_ranks=0)
al.dfs_to_ws(ws1,1,1,df_MC0,hd=True)

# 将wb封装为一个二进行的BytesIO文件.
down_file=wb_to_bytesIO(wb)

# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# C:添加一个下载按钮
# ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 添加时间信息。
st.success(f"运算己经完成，共用时：{round(time.time() - start, 2)}秒。")
# 创建下载按钮,以便下载此工作簿
st.download_button(
    label='下载分析结果',
    data=down_file,
    file_name=os.path.splitext(uploaded_file.name)[0] + "_" + sht_MB_name + "_报表" + ".xlsx",
    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')





